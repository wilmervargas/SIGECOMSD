
import os
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa
from django.conf import settings

# SEGURIDAD HISTORICO DE LOS REGISTROS
from django.contrib.contenttypes.models import ContentType # Para identificar el modelo
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION      # LogEntry y Flags de acción
# 💡 NUEVA IMPORTACIÓN PARA MANEJO DE ERRORES DE BASE DE DATOS
from django.db import IntegrityError 
# 💡 NUEVA IMPORTACIÓN PARA EL REGISTRO DE EVENTOS (LOGGING)
import logging
from django.db.models import ProtectedError # 👈 Importar el error específico

# -------------------------------------------------------------

# Importa el formulario de dependencias
from dependencias.forms import DependenciasFilterForm, DependenciasForm
from dependencias.models import DependenciasBD

from django.db.models import Q

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage
from django.http import Http404, HttpResponse
from django.urls import reverse
from datetime import datetime

logger = logging.getLogger(__name__)

@login_required
def listado_dependencias(request):

    # 🛑 Usa el formulario de filtro simple aquí
    filter_form = DependenciasFilterForm(request.GET)
    
    # 🛑 Asegúrate de que tu queryset base sea DependenciasBD.objects.all() 
    queryset_dependencias = DependenciasBD.objects.all() 

    # =========================================================
    # 📌 CÓDIGO AÑADIDO/MODIFICADO PARA PERSISTENCIA DE FILTROS
    # =========================================================
    
    # 1. Copiamos todos los parámetros GET.
    query_params = request.GET.copy()
    
    # 2. Eliminamos el parámetro 'page' para que no se duplique en los enlaces.
    if 'page' in query_params:
        del query_params['page']
        
    # 3. Codificamos los filtros y ordenación restantes en una cadena (search_query, order_by, etc.).
    query_string = query_params.urlencode()
    # =========================================================
    clean_params = query_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    
    # Esta variable contendrá solo los filtros (ej. 'search_query=abc')
    clean_query_string = clean_params.urlencode()
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # FILTRO 1: Búsqueda por Código o Descripción
        search_query = data.get('search_query')
        if search_query:
            # Tu lógica de filtrado es correcta:
            queryset_dependencias = queryset_dependencias.filter(
                Q(cod_dependencia__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query))
        else:
            print("El formulario NO es válido. Errores:", filter_form.errors) # <-- Línea de depuración 3

    # 2. ORDENACIÓN (Simplificada)
    # Obtenemos el parámetro de ordenación (por defecto 'cod_dependencia')
    # Usamos 'cod_dependencia' como desempate si la ordenación principal es 'descripcion'
    order_by_principal = request.GET.get('order_by', 'cod_dependencia') 
    
    # 💡 Lógica de desempate simple: siempre usamos cod_dependencia si no es el principal
    if order_by_principal.lstrip('-') == 'descripcion':
        # Si ordenamos por descripción, el desempate es por cod_dependencia
        lista_dependencias = queryset_dependencias.order_by(order_by_principal, 'cod_dependencia')
    else:
        # Si ordenamos por cod_dependencia (o cualquier otro campo), no necesitamos desempate
        lista_dependencias = queryset_dependencias.order_by(order_by_principal)

    # --- Paginación ---
    page = request.GET.get('page', 1)
    try:
        paginator = Paginator(lista_dependencias, 50)
        lista_dependencias = paginator.page(page)
    except EmptyPage:
        # Si se accede a una página fuera de rango, ir a la última página
        lista_dependencias = paginator.page(paginator.num_pages)
    except:
        # Cualquier otro error (ej. texto no numérico en 'page')
        raise Http404('*** Página no encontrada ***')
    # --- Paginación ---

    # --- Contexto ---
    datos = {
        'titulo': 'dependencias',
        'subpagina': 'subpage', 
        'entity': lista_dependencias, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        'query_string': query_string, # ⬅️ LÍNEA AÑADIDA
        'clean_query_string': clean_query_string, 
    }
    return render(request, 'dependencias/listado.html', datos)

# =================================================================
# VISTAS CRUD (Crear, Editar, Borrar) - Variables Adaptadas
# =================================================================
@login_required
def crear_dependencias(request):

    formulario = DependenciasForm(request.POST or None) 
    datos = {'titulo': 'Tabla dependencias', 'subpagina': 'subpage', 'formulario': formulario,}
    order_by_param = request.GET.get('order_by', '')
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Contiene el código que debe ser monitoreado
            
            # 1. Guarda y obtén el objeto Producto creado
            dependencias = formulario.save() 
            
            # 🟢 REGISTRO DE LOG (Creación Exitosa) --------------------------
            LogEntry.objects.create(
                        user_id=request.user.pk,
                        content_type_id=ContentType.objects.get_for_model(dependencias).pk,
                        object_id=dependencias.pk,
                        object_repr=str(dependencias),
                        action_flag=DELETION, # Bandera para Eliminación (3)
                        change_message=f'Dependencia {dependencias.cod_dependencia} creado mediante vista personalizada.'
                    )
        
            messages.success(request, f'Dependencia {dependencias.cod_dependencia} fue creado exitosamente', extra_tags='procesado ✅')

            base_url = reverse('listado_dependencias')
            return redirect(f'{base_url}?order_by={order_by_param}') if order_by_param else redirect('listado_dependencias')
        
        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores (Overflow, DB o cualquier otro)
            
            # 1. Registro detallado en el Log del Sistema (Recomendado)
            # exc_info=True asegura que se grabe el StackTrace (la pila de llamadas)
            logger.error(f"Fallo crítico al crear Dependencia para usuario {request.user.pk}: {e}", exc_info=True)
            
            # 2. Registro de Fallo en la tabla LogEntry (Para el historial del Admin)
            # Usamos 4 (o el valor que definas para ERROR) ya que ADDITION, CHANGE y DELETION son 1, 2 y 3.
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(DependenciasBD).pk, 
                object_id=None, # No hay ID de objeto porque la creación falló
                object_repr='Fallo en la creación de Dependencia', 
                action_flag=4, # Usamos '4' (o un número fuera de ADDITION/CHANGE/DELETION) para Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo crear el Dependencia. Causa: {type(e).__name__}. Ver log del servidor.'
            )
            # 3. Muestra un mensaje amigable al usuario
            messages.error(request, 'Error grave de ejecución. El Dependencia NO fue creado. Por favor, contacte a soporte.', extra_tags='error ❌')

    return render (request, 'dependencias/crear.html', datos)


#------------------------------------------
@login_required
def editar_dependencias(request, id):
#------------------------------------------
    dependencias = get_object_or_404(DependenciasBD, id=id) 
    order_by_param = request.GET.get('order_by', '')
    formulario = DependenciasForm(request.POST or None, instance=dependencias)
    
    datos = {
        'titulo': 'Tabla dependencias', 'subpagina': 'subpage', 'formulario': formulario, 'order_by': order_by_param,
    }
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Contiene el código que debe ser monitoreado
            
            # 1. Capturar la lista de campos modificados antes de guardar
            campos_modificados = formulario.changed_data 
            
            # 2. Guardar los cambios (puede fallar por IntegrityError)
            dependencias = formulario.save() 
            
            # 3. Construir el mensaje de cambio
            if campos_modificados:
                change_message = 'Campos modificados: ' + ', '.join(campos_modificados)
            else:
                change_message = 'Dependencia fue guardada, pero no se detectaron cambios.'
                
            # 🟡 REGISTRO DE LOG (Modificación Exitosa) ------------------------
            LogEntry.objects.create(
                        user_id=request.user.pk,
                        content_type_id=ContentType.objects.get_for_model(dependencias).pk,
                        object_id=dependencias.pk,
                        object_repr=str(dependencias),
                        action_flag=DELETION, # Bandera para Eliminación (3)
                        change_message=f'Dependencia {dependencias.cod_dependencia} editada mediante vista personalizada.'
                    )
            # --------------------------------------------------------

            messages.success(request, f'Dependencia {dependencias.cod_dependencia} editada exitosamente', extra_tags='procesado ✅')
            base_url = reverse('listado_dependencias')
            return redirect(f'{base_url}?order_by={order_by_param}') if order_by_param else redirect('listado_dependencias')

        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores durante la edición
            
            # 1. Registro detallado en el Log del Sistema
            logger.error(f"Fallo crítico al editar catgoría ID {id} para usuario {request.user.pk}: {e}", exc_info=True)
            
            # 2. Registro de Fallo en la tabla LogEntry (Admin History)
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(DependenciasBD).pk, 
                object_id=None, # No hay ID de objeto porque la creación falló
                object_repr='Fallo en la edición de Dependencia',
                action_flag=4, # Usamos '4' (o un número fuera de ADDITION/CHANGE/DELETION) para Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo editar el Dependencia. Causa: {type(e).__name__}. Ver log del servidor.'
            )

            # 3. Muestra un mensaje amigable al usuario
            messages.error(request, 'Error grave de ejecución al editar. Los cambios NO fueron guardados. Contacte a soporte.', extra_tags='error ❌')

    return render(request, 'dependencias/editar.html', datos)

#------------------------------------------
@login_required
def borrar_dependencias(request, id):
    # 1. Obtener la instancia de la dependencia o devolver 404
    # Basado en tu modelo DependenciasBD
    dependencia = get_object_or_404(DependenciasBD, id=id) 
    
    id_dependencia = dependencia.cod_dependencia 
    
    # =========================================================
    # 📌 PASO 1: CAPTURAR PARÁMETROS GET 
    # Mantiene la persistencia de filtros, orden y paginación.
    # =========================================================
    query_string = request.GET.urlencode()

    try:
        # 2. Intentar la eliminación
        # Esto disparará el ProtectedError debido a 'on_delete=models.PROTECT' 
        # en el modelo DependeciasDirectorBD.
        dependencia.delete()
        
        # 3. Registro en Log y mensaje de éxito
        LogEntry.objects.create(
            user_id=request.user.pk,
            content_type_id=ContentType.objects.get_for_model(dependencia).pk,
            object_id=dependencia.pk,
            object_repr=str(dependencia),
            action_flag=DELETION,
            change_message=f'Dependencia {id_dependencia} eliminada correctamente.'
        )
        
        messages.success(request, f'La Dependencia {id_dependencia} fue borrada exitosamente.', extra_tags='procesado ✅')
        
    except ProtectedError as e:
        # 4. CAPTURAR la protección (si hay directores vinculados)
        logger.error(f"Intento de borrado bloqueado: Dependencia {id_dependencia} está en uso.")
        
        # Analizamos qué modelo está bloqueando la eliminación
        if e.args and len(e.args) > 1 and isinstance(e.args[1], set):
            # Obtenemos el nombre del modelo que causó el bloqueo (ej. DependeciasDirectorBD)
            referencing_objects = e.args[1]
            model_name = next(iter(referencing_objects)).__class__.__name__
            
            # Personalizamos el mensaje según el modelo detectado
            if model_name == 'DependeciasDirectorBD':
                msg = (
                    f"ERROR: No se puede eliminar la Dependencia {id_dependencia}. "
                    "Existen directores o encargados asociados a esta oficina. "
                    "Debe desvincularlos en el módulo de Directores antes de proceder."
                )
            else:
                msg = f"ERROR: La Dependencia {id_dependencia} está enlazada al modelo '{model_name}'."
        else:
            msg = f"ERROR: No se puede borrar la dependencia {id_dependencia} porque otros registros dependen de ella."

        messages.error(request, msg, extra_tags='error ❌')
        
    # 5. Redirección persistente
    base_url = reverse('listado_dependencias')
    if query_string:
        return redirect(f'{base_url}?{query_string}')
    return redirect('listado_dependencias')


# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE EXCEL -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_dependencias_excel(request):

    # --- CAMBIO PARA FILTRO Y ORDEN ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'cod_dependencia')

    dependencias = DependenciasBD.objects.all()
    if search_query:
        dependencias = dependencias.filter(
            Q(cod_dependencia__icontains=search_query) | 
            Q(observaciones__icontains=search_query) | 
            Q(descripcion__icontains=search_query)
        )
    dependencias = dependencias.order_by(order_by)
    # ----------------------------------

    # 1. Configuración Inicial y Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Dependencias_Formato.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Dependencias"
    
    # --- 2. INSERCIÓN DEL LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        ws.row_dimensions[1].height = 50
        ws.column_dimensions['A'].width = 15 
        img.anchor = 'A1' 
        ws.add_image(img)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el logo en la ruta: {logo_path}")
    
    # --- 3. CREACIÓN Y ESTILOS DEL ENCABEZADO (Título) ---
    ws.merge_cells('B2:D2')
    title_font = Font(name='Arial', size=14, bold=True, color="1F497D") 
    title_alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'] = "REPORTE DE DEPENDENCIAS"
    ws['B2'].font = title_font
    ws['B2'].alignment = title_alignment
    
    # --- 4. ENCABEZADOS DE DATOS Y ESTILOS DE TABLA ---
    ws.append(()) 
    header_font = Font(bold=True, color="FFFFFF") 
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    data_font = Font(size=10, name='Arial')
    center_alignment = Alignment(horizontal='center', vertical='center') 
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    
    fill_even_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    fill_odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    start_row = 4
    headers = ['Nro.', 'Código', 'Descripción Dependencia', 'Observaciones']
    ws.append(headers) 

    for cell in ws[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # --- 5. DATOS DEL MODELO ---
    data_start_row = start_row + 1 
    
    # SE ELIMINÓ LA LÍNEA QUE SOBREESCRIBÍA LAS depenORÍAS FILTRADAS
    for i, depen in enumerate(dependencias, 1): 
        row_data = [i, depen.cod_dependencia, depen.descripcion, depen.observaciones]
        ws.append(row_data)

        current_row_index = data_start_row + i - 1
        current_row = ws[current_row_index]
        ws.row_dimensions[current_row_index].height = 40
        row_fill = fill_even_row if i % 2 == 0 else fill_odd_row
        
        for col_index, cell in enumerate(current_row):
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill 
            if col_index in [0, 1]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # --- 6. AUTOAJUSTE Y LIMITACIÓN DE COLUMNAS ---
    column_widths = {} 
    MAX_WIDTH_C = 40 
    MAX_WIDTH_D = 40 

    for row in ws.iter_rows(min_row=start_row):
        for i, cell in enumerate(row):
            col_letter = get_column_letter(i + 1)
            try:
                if cell.value is not None:
                    length = max(10, len(str(cell.value))) 
                    if i == 2: length = min(length, MAX_WIDTH_C)
                    elif i == 3: length = min(length, MAX_WIDTH_D)
                    current_max = column_widths.get(col_letter, 0)
                    if length > current_max:
                        column_widths[col_letter] = length
            except:
                pass 

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width + 3 
        
    wb.save(response)
    return response


# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE PDF -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_dependencias_pdf(request):
    
    # --- 1. CAPTURA DE FILTRO Y ORDEN (Mantiene la persistencia) ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'descripcion')

    try:
        # Iniciamos el queryset
        dependencias = DependenciasBD.objects.all()
        
        # Aplicamos filtro si existe búsqueda
        if search_query:
            dependencias = dependencias.filter(
                Q(cod_dependencia__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )
        
        # Aplicamos el orden seleccionado
        dependencias = dependencias.order_by(order_by)
        
    except Exception as e:
        # En caso de error de base de datos o modelo no importado
        print(f"Error en reporte: {e}")
        dependencias = []

    # --- 2. Preparación del Contexto para la Plantilla PDF ---
    
    # IMPORTANTE: Generar la URL absoluta para el logo
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except AttributeError:
        logo_url = "" 

    # Obtener la fecha y hora actual
    fecha_actual = datetime.now() 
    
    context = {
        'dependencias': dependencias, # 👈 Esta variable ya viene filtrada y ordenada
        'titulo_reporte': 'Listado de Dependencias',
        'logo_path': logo_url,
        'fecha_emision': fecha_actual,
    }

    # --- 3. Renderizado y Conversión a PDF ---
    template = get_template('reportes/reporte_dependencias_pdf.html')
    html = template.render(context)

    result = BytesIO()
    
    pisa_status = pisa.CreatePDF(
       html,              
       dest=result,       
       encoding='utf-8',  
       link_callback=lambda uri, rel: uri
    )

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: %s' % pisa_status.err, status=500)
    
    # --- 4. Devolver la Respuesta con Apertura en Nueva Pestaña ---
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_dependencias.pdf"'
    
    return response