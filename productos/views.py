
import os
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa
from django.conf import settings

# SEGURIDAD HISTORICO DE LOS REGISTROS
from django.contrib.contenttypes.models import ContentType # Para identificar el modelo
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION      # LogEntry y Flags de acción
# 💡 NUEVA IMPORTACIÓN PARA MANEJO DE ERRORES DE BASE DE DATOS
from django.db import IntegrityError 
# 💡 NUEVA IMPORTACIÓN PARA EL REGISTRO DE EVENTOS (LOGGING)
import logging
from django.db.models import ProtectedError # 👈 Importar el error específico

# Importa el formulario de productos
from productos.forms import ProductosFilterForm, ProductosForm
from productos.models import Producto
from django.db.models import Q, F  # <--- Agrega F aquí
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage
from django.http import Http404, HttpResponse
from django.urls import reverse
from datetime import datetime

logger = logging.getLogger(__name__)
from django.contrib.messages import get_messages
@login_required
def listado_productos(request):

    # Esto consume todos los mensajes acumulados y los deja vacíos
    storage = get_messages(request)
    for message in storage:
        pass  # Solo iterar para "limpiarlos" del storage

    # 1. Inicialización del formulario con los datos GET
    filter_form = ProductosFilterForm(request.GET or None)
    queryset_productos = Producto.objects.all() 

    # --- Lógica de persistencia de filtros (Tu estructura original) ---
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()
    
    clean_params = query_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    clean_query_string = clean_params.urlencode()

    # --- APLICACIÓN DE FILTROS ---
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # 1. Búsqueda por texto (Código, Observaciones, Descripción)
        search_query = data.get('search_query')
        if search_query:
            queryset_productos = queryset_productos.filter(
                Q(cod_producto__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )

        # 2. Filtro por Categoría
        if data.get('categoria'):
            queryset_productos = queryset_productos.filter(cod_categoria=data.get('categoria'))

        # 3. Filtro por Unidad
        if data.get('unidad'):
            queryset_productos = queryset_productos.filter(cod_unidad=data.get('unidad'))

        # 4. Filtro por Alerta de Stock (Comparación entre campos)
        alerta = data.get('alerta_stock')
        if alerta == 'MIN':
            # Muestra productos donde el stock actual es menor o igual al mínimo global
            queryset_productos = queryset_productos.filter(cantidad__lte=F('stock_minimo_global'))
        elif alerta == 'MAX':
            # Muestra productos con stock normal
            queryset_productos = queryset_productos.filter(cantidad__gt=F('stock_minimo_global'))

        # 5. Filtro de Status (Campo lactivo)
        lactivo = data.get('lactivo')
        if lactivo == 'True':
            queryset_productos = queryset_productos.filter(activo=True)
        elif lactivo == 'False':
            queryset_productos = queryset_productos.filter(activo=False)
        # Si es '', muestra todos (no se aplica filtro)
    else:
        # Si el form no es válido al inicio (o no hay GET), filtramos por Activos por defecto
        # para cumplir con "siempre mantener a la vista los activos"
        queryset_productos = queryset_productos.filter(activo=True)

    # --- ORDENACIÓN ---
    order_by_principal = request.GET.get('order_by', 'cod_producto')

    # 1. Ordenar por descripción de Categoría
    if order_by_principal == 'cod_categoria':
        lista_productos = queryset_productos.order_by('cod_categoria__descripcion', 'cod_producto')
    elif order_by_principal == '-cod_categoria':
        lista_productos = queryset_productos.order_by('-cod_categoria__descripcion', 'cod_producto')

    # 2. Ordenar por descripción de Unidad (NUEVO)
    elif order_by_principal == 'cod_unidad':
        lista_productos = queryset_productos.order_by('cod_unidad__descripcion', 'cod_producto')
    elif order_by_principal == '-cod_unidad':
        lista_productos = queryset_productos.order_by('-cod_unidad__descripcion', 'cod_producto')

    # 3. Lógica existente para descripción y otros campos
    elif order_by_principal.lstrip('-') == 'descripcion':
        lista_productos = queryset_productos.order_by(order_by_principal, 'cod_producto')
    else:
        lista_productos = queryset_productos.order_by(order_by_principal)

    # --- PAGINACIÓN ---
    page = request.GET.get('page', 1)
    paginator = Paginator(lista_productos, 50)
    try:
        lista_productos = paginator.page(page)
    except EmptyPage:
        lista_productos = paginator.page(paginator.num_pages)
    except:
        raise Http404('*** Página no encontrada ***')

    # --- CONTEXTO (Mantiene tus variables: entity, filter_form, etc.) ---
    datos = {
        'titulo': 'insumos',
        'subpagina': 'subpage', 
        'entity': lista_productos, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        'query_string': query_string, 
        'clean_query_string': clean_query_string, 
    }
    return render(request, 'productos/listado.html', datos)

# =================================================================
# VISTAS CRUD (Crear, Editar, Borrar) - Variables Adaptadas
# =================================================================
@login_required
def crear_productos(request):

    formulario = ProductosForm(request.POST or None, request.FILES or None) # 🟢 Añade request.FILES para la imagen
    datos = {'titulo': 'Tabla Productos', 'subpagina': 'subpage', 'formulario': formulario,}
    order_by_param = request.GET.get('order_by', '')
    
    if formulario.is_valid():
        try:
            productos = formulario.save() 
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(productos).pk,
                object_id=productos.pk,
                object_repr=str(productos),
                action_flag=ADDITION, # 🟢 Cambiado a ADDITION
                change_message=f'Producto {productos.cod_producto} creado.'
            )

            messages.success(request, f'Insumo {productos.cod_producto} fue creado exitosamente', extra_tags='procesado ✅')

            base_url = reverse('listado_productos')
            return redirect(f'{base_url}?order_by={order_by_param}') if order_by_param else redirect('listado_productos')
        
        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores (Overflow, DB o cualquier otro)
            
            # 1. Registro detallado en el Log del Sistema (Recomendado)
            # exc_info=True asegura que se grabe el StackTrace (la pila de llamadas)
            logger.error(f"Fallo crítico al crear Producto para usuario {request.user.pk}: {e}", exc_info=True)
            
            # 2. Registro de Fallo en la tabla LogEntry (Para el historial del Admin)
            # Usamos 4 (o el valor que definas para ERROR) ya que ADDITION, CHANGE y DELETION son 1, 2 y 3.
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(Producto).pk, 
                object_id=None, # No hay ID de objeto porque la creación falló
                object_repr='Fallo en la creación de producto', 
                action_flag=4, # Usamos '4' (o un número fuera de ADDITION/CHANGE/DELETION) para Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo crear el Insumo. Causa: {type(e).__name__}. Ver log del servidor.'
            )
            # 3. Muestra un mensaje amigable al usuario
            messages.error(request, 'Error grave de ejecución. El Insumo NO fue creado. Por favor, contacte a soporte.', extra_tags='error ❌')
            print(f"Errores del formulario: {formulario.errors}")
            
    return render (request, 'productos/crear.html', datos)


#------------------------------------------
# ------------------------------------------
@login_required
def editar_productos(request, id):
# ------------------------------------------
    productos = get_object_or_404(Producto, id=id) 
    
    # 1. CAPTURAR TODA LA QUERY (Filtros + Orden + Página)
    # Esto guarda 'search_query=abc&categoria=1&order_by=-precio' etc.
    query_string = request.GET.urlencode()
    
    # 2. Inicializar el formulario con los datos del POST (si existen) y la instancia del producto
    formulario = ProductosForm(request.POST or None, request.FILES or None, instance=productos)
    
    if formulario.is_valid():
        try:
            # Capturar campos modificados para el log
            campos_modificados = formulario.changed_data 
            
            # Guardar los cambios en la base de datos
            productos = formulario.save() 
            
            # Construir mensaje para el historial
            if campos_modificados:
                change_message = f'Campos modificados: {", ".join(campos_modificados)}'
            else:
                change_message = 'Se guardó sin detectar cambios en los campos.'
                
            # REGISTRO DE LOG (Modificación Exitosa)
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(productos).pk,
                object_id=productos.pk,
                object_repr=str(productos),
                action_flag=CHANGE, # Flag 2 para ediciones
                change_message=f'Insumo {productos.cod_producto} editado. {change_message}'
            )

            messages.success(request, f'Insumo {productos.cod_producto} editado exitosamente', extra_tags='procesado ✅')
            
            # 3. REDIRECCIÓN INTELIGENTE
            # Reconstruimos la URL de retorno con todos los parámetros originales
            base_url = reverse('listado_productos')
            if query_string:
                return redirect(f'{base_url}?{query_string}')
            return redirect('listado_productos')

        except (OverflowError, IntegrityError, Exception) as e:
            # Registro de error en el log del servidor
            logger.error(f"Fallo crítico al editar insumo ID {id}: {e}", exc_info=True)
            
            # Registro de fallo en LogEntry para el Admin
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(Producto).pk, 
                object_id=productos.pk,
                object_repr=f'Error en edición: {productos.cod_producto}',
                action_flag=4, 
                change_message=f'ERROR: No se pudo editar. Causa: {type(e).__name__}'
            )

            messages.error(request, 'Error grave al editar. Los cambios NO se guardaron.', extra_tags='error ❌')

    # 4. PASAR LA QUERY AL CONTEXTO
    # Es vital pasar 'query_string' para que el botón "Cancelar" también mantenga los filtros
    datos = {
        'titulo': 'Tabla Insumos', 
        'subpagina': 'subpage', 
        'formulario': formulario, 
        'query_string': query_string, # <--- IMPORTANTE
    }
    
    return render(request, 'productos/editar.html', datos)
        
#------------------------------------------
@login_required
def borrar_productos(request, id):
# 1. Obtener la instancia del objeto o devolver 404
    # Asume que tu modelo de Producto es Producto
    productos = get_object_or_404(Producto, id=id) 
    
    # Asume que el código de la Producto es 'cod_producto'
    id_productos = productos.cod_producto 
    
    # =========================================================
    # 📌 PASO 1: CAPTURAR TODOS LOS PARÁMETROS GET 
    # Esto captura y mantiene el filtro, orden y página.
    # =========================================================
    query_string = request.GET.urlencode()

    try:
        # 2. Intentar la eliminación (puede causar ProtectedError)
        productos.delete()
        
        # 3. Si es exitoso, registrar la acción y mostrar el mensaje de éxito.
        # 🔴 REGISTRO DE LOG (Eliminación Exitosa) --------------------------
        LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(productos).pk,
                    object_id=productos.pk,
                    object_repr=str(productos),
                    action_flag=DELETION, # Bandera para Eliminación (3)
                    change_message=f'Insumo {id_productos} eliminado mediante vista personalizada.'
                )
        
        messages.success(request, f'Insumo {id_productos} fue borrado exitosamente', extra_tags='procesado ✅')
        
    except ProtectedError as e:
        # 4. CAPTURAR el ProtectedError (el registro está enlazado)
        logger.error(f"Intento fallido de borrar Insumo {id_productos}")       
        
        # Generar mensaje de error amigable al usuario
        if e.args and isinstance(e.args[1], set) and e.args[1]:
            # Extraemos el nombre del modelo relacionado que impide la eliminación
            referencing_model = next(iter(e.args[1])).__class__.__name__
            msg = (
                f"ERROR: No se puede eliminar el Insumo {id_productos}. "
                f"Está enlazada a registros en el modelo **'{referencing_model}'**  "
                f"Por favor, elimine o reasigne los registros enlazados primero."
            )
        else:
            msg = (
                f"ERROR: No se puede eliminar el Insumo {id_productos} porque "
                f"está siendo referenciada por otros registros. "
                f"Debe eliminarlos o reasignarlos primero."
            )

        messages.error(request, msg, extra_tags='error ❌')
        
    # 5. Redirigir a la lista de productos, ADJUNTANDO todos los parámetros GET.
    # Esto asegura que mantenga el filtro, orden y página.
    base_url = reverse('listado_productos') # Asume que tu URL de listado se llama 'listado_productos'
    
    if query_string:
        # Si hay parámetros (filtros, etc.), los adjuntamos con '?'
        return redirect(f'{base_url}?{query_string}')
    else:
        # Si no había parámetros, redirigimos a la base.
        return redirect('listado_productos')

# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE EXCEL -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_productos_excel(request):
    # --- 1. CAPTURA DE FILTROS (Sincronizado con PDF) ---
    filter_form = ProductosFilterForm(request.GET or None)
    productos = Producto.objects.all()

    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        search_query = data.get('search_query')
        if search_query:
            productos = productos.filter(
                Q(cod_producto__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )

        if data.get('categoria'):
            productos = productos.filter(cod_categoria=data.get('categoria'))

        if data.get('unidad'):
            productos = productos.filter(cod_unidad=data.get('unidad'))

        alerta = data.get('alerta_stock')
        if alerta == 'MIN':
            productos = productos.filter(cantidad__lte=F('stock_minimo_global'))
        elif alerta == 'MAX':
            productos = productos.filter(cantidad__gt=F('stock_minimo_global'))

        lactivo = data.get('lactivo')
        if lactivo == 'True':
            productos = productos.filter(activo=True)
        elif lactivo == 'False':
            productos = productos.filter(activo=False)
    else:
        productos = productos.filter(activo=True)

       # --- ORDENACIÓN ---
    order_by = request.GET.get('order_by', 'cod_producto')

    # 1. Ordenar por descripción de Categoría
    if order_by == 'cod_categoria':
        productos = productos.order_by('cod_categoria__descripcion', 'cod_producto')
    elif order_by == '-cod_categoria':
        productos = productos.order_by('-cod_categoria__descripcion', 'cod_producto')

    # 2. Ordenar por descripción de Unidad (NUEVO)
    elif order_by == 'cod_unidad':
        productos = productos.order_by('cod_unidad__descripcion', 'cod_producto')
    elif order_by == '-cod_unidad':
        productos = productos.order_by('-cod_unidad__descripcion', 'cod_producto')

    # 3. Lógica existente para descripción y otros campos
    elif order_by.lstrip('-') == 'descripcion':
        productos = productos.order_by(order_by, 'cod_producto')
    else:
        productos = productos.order_by(order_by)

    # --- 3. CONFIGURACIÓN EXCEL ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Insumos.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"
    
    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        img.width = 90
        img.height = 40
        ws.add_image(img, 'A1')
    except:
        pass
    
    # Título Principal
    ws.merge_cells('B3:J3')
    ws['B3'] = "REPORTE DETALLADO DE INSUMOS"
    ws['B3'].font = Font(name='Arial', size=14, bold=True, color="1F497D")
    ws['B3'].alignment = Alignment(horizontal='center')
    
    # --- 4. ENCABEZADOS DE TABLA (Basados en el HTML) ---
    start_row = 4
    headers = [
        'Nro', 'Código', 'Descripción', 'Categoría', 
        'Unidad', 'Stock Mín', 'Stock Máx', 'Existencia', 
        'Costo', 'Observaciones'
    ]
    
    header_fill = PatternFill(start_color="1a4d94", end_color="1a4d94", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append([]) # Fila vacía para espacio
    ws.append(headers)

    for cell in ws[start_row + 1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- 5. DATOS ---
    fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for i, prod in enumerate(productos, 1):
        row_data = [
            i,
            prod.cod_producto,
            prod.descripcion.upper(),
            str(prod.cod_categoria.descripcion) if prod.cod_categoria else "-",
            str(prod.cod_unidad.descripcion) if prod.cod_unidad else "-",
            prod.stock_minimo_global,
            prod.stock_maximo,
            prod.cantidad if prod.cantidad is not None else "0,00",
            prod.costo_compra,
            (prod.observaciones or "").upper() or "-"
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx, cell in enumerate(ws[curr_row], 1):
            cell.border = thin_border
            cell.font = Font(size=9)
            if i % 2 == 0: cell.fill = fill_even
            
            # Alineación y formatos numéricos
            if col_idx in [1, 2, 4, 5]: # Nro, Código, Cat, Unidad
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in [6, 7, 8, 9]: # Valores numéricos
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
            else: # Descripción y Observaciones
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

    # --- 6. AUTOAJUSTE ---
    # Ajustamos anchos específicos
    dims = {'A': 5, 'B': 15, 'C': 40, 'D': 20, 'E': 15, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 40}
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save(response)
    return response


# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE PDF -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_productos_pdf(request):
    # --- 1. CAPTURA DE FILTROS (Usando el mismo Form que el listado) ---
    filter_form = ProductosFilterForm(request.GET or None)
    productos = Producto.objects.all()

    # --- 2. APLICACIÓN DE LA LÓGICA DE FILTROS ---
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # Búsqueda por texto
        search_query = data.get('search_query')
        if search_query:
            productos = productos.filter(
                Q(cod_producto__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )

        # Filtro por Categoría
        if data.get('categoria'):
            productos = productos.filter(cod_categoria=data.get('categoria'))

        # Filtro por Unidad
        if data.get('unidad'):
            productos = productos.filter(cod_unidad=data.get('unidad'))

        # Filtro por Alerta de Stock
        alerta = data.get('alerta_stock')
        if alerta == 'MIN':
            productos = productos.filter(cantidad__lte=F('stock_minimo_global'))
        elif alerta == 'MAX':
            productos = productos.filter(cantidad__gt=F('stock_minimo_global'))

        # Filtro de Status (Campo activo)
        lactivo = data.get('lactivo')
        if lactivo == 'True':
            productos = productos.filter(activo=True)
        elif lactivo == 'False':
            productos = productos.filter(activo=False)
    else:
        # Si no hay filtros válidos, por defecto mostramos solo activos
        productos = productos.filter(activo=True)

    # --- ORDENACIÓN ---
    order_by = request.GET.get('order_by', 'cod_producto')

    # 1. Ordenar por descripción de Categoría
    if order_by == 'cod_categoria':
        productos = productos.order_by('cod_categoria__descripcion', 'cod_producto')
    elif order_by == '-cod_categoria':
        productos = productos.order_by('-cod_categoria__descripcion', 'cod_producto')

    # 2. Ordenar por descripción de Unidad (NUEVO)
    elif order_by == 'cod_unidad':
        productos = productos.order_by('cod_unidad__descripcion', 'cod_producto')
    elif order_by == '-cod_unidad':
        productos = productos.order_by('-cod_unidad__descripcion', 'cod_producto')

    # 3. Lógica existente para descripción y otros campos
    elif order_by.lstrip('-') == 'descripcion':
        productos = productos.order_by(order_by, 'cod_producto')
    else:
        productos = productos.order_by(order_by)


    # --- 4. PREPARACIÓN DEL CONTEXTO ---
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except:
        logo_url = "" 

    context = {
        'productos': productos,
        'titulo_reporte': 'Reporte Detallado de Insumos',
        'logo_path': logo_url,
        'fecha_emision': datetime.now(),
    }

    # --- 5. RENDERIZADO Y GENERACIÓN DE PDF ---
    template = get_template('reportes/reporte_productos_pdf.html')
    html = template.render(context)
    result = BytesIO()
    
    pisa_status = pisa.CreatePDF(
       html,              
       dest=result,       
       encoding='utf-8',  
       link_callback=lambda uri, rel: uri
    )

    if pisa_status.err:
        return HttpResponse(f'Error al generar el PDF: {pisa_status.err}', status=500)
    
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_productos.pdf"'
    
    return response

