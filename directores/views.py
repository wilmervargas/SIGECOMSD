
import os
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa
from django.conf import settings

# SEGURIDAD HISTORICO DE LOS REGISTROS
from django.contrib.contenttypes.models import ContentType # Para identificar el modelo
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION      # LogEntry y Flags de acción
# 💡 NUEVA IMPORTACIÓN PARA MANEJO DE ERRORES DE BASE DE DATOS
from django.db import IntegrityError 
# 💡 NUEVA IMPORTACIÓN PARA EL REGISTRO DE EVENTOS (LOGGING)
import logging
from django.db.models import ProtectedError # 👈 Importar el error específico

# -------------------------------------------------------------

# Importa el formulario de directores
from directores.forms import DirectoresFilterForm, DirectoresForm, DependenciasFormSet
from directores.models import DirectoresBD

from django.db.models import Q

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage
from django.http import Http404, HttpResponse
from django.urls import reverse
from datetime import datetime

logger = logging.getLogger(__name__)

@login_required
def listado_directores(request):

    # 🛑 Usa el formulario de filtro simple aquí
    filter_form = DirectoresFilterForm(request.GET)
    
    # 🛑 Asegúrate de que tu queryset base sea DirectoresBD.objects.all() 
    queryset_directores = DirectoresBD.objects.all() 

    # =========================================================
    # 📌 CÓDIGO AÑADIDO/MODIFICADO PARA PERSISTENCIA DE FILTROS
    # =========================================================
    
    # 1. Copiamos todos los parámetros GET.
    query_params = request.GET.copy()
    
    # 2. Eliminamos el parámetro 'page' para que no se duplique en los enlaces.
    if 'page' in query_params:
        del query_params['page']
        
    # 3. Codificamos los filtros y ordenación restantes en una cadena (search_query, order_by, etc.).
    query_string = query_params.urlencode()
    # =========================================================
    clean_params = query_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    
    # Esta variable contendrá solo los filtros (ej. 'search_query=abc')
    clean_query_string = clean_params.urlencode()
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # FILTRO 1: Búsqueda por Código o Descripción
        search_query = data.get('search_query')
        if search_query:
            # Tu lógica de filtrado es correcta:
            queryset_directores = queryset_directores.filter(
                Q(cedula__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(usuario__icontains=search_query) | 
                Q(nombres_apellidos__icontains=search_query))
        else:
            print("El formulario NO es válido. Errores:", filter_form.errors) # <-- Línea de depuración 3

    # 2. ORDENACIÓN (Simplificada)
    # Obtenemos el parámetro de ordenación (por defecto 'cedula')
    # Usamos 'cedula' como desempate si la ordenación principal es 'nombres_apellidos'
    order_by_principal = request.GET.get('order_by', 'cedula') 
    
    # 💡 Lógica de desempate simple: siempre usamos cedula si no es el principal
    if order_by_principal.lstrip('-') == 'nombres_apellidos':
        # Si ordenamos por descripción, el desempate es por cedula
        lista_directores = queryset_directores.order_by(order_by_principal, 'cedula')
    else:
        # Si ordenamos por cedula (o cualquier otro campo), no necesitamos desempate
        lista_directores = queryset_directores.order_by(order_by_principal)

    # --- Paginación ---
    page = request.GET.get('page', 1)
    try:
        paginator = Paginator(lista_directores, 50)
        lista_directores = paginator.page(page)
    except EmptyPage:
        # Si se accede a una página fuera de rango, ir a la última página
        lista_directores = paginator.page(paginator.num_pages)
    except:
        # Cualquier otro error (ej. texto no numérico en 'page')
        raise Http404('*** Página no encontrada ***')
    # --- Paginación ---

    # --- Contexto ---
    datos = {
        'titulo': 'directores',
        'subpagina': 'subpage', 
        'entity': lista_directores, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        'query_string': query_string, # ⬅️ LÍNEA AÑADIDA
        'clean_query_string': clean_query_string, 
    }
    return render(request, 'directores/listado.html', datos)

# =================================================================
# VISTAS CRUD (Crear, Editar, Borrar) - Variables Adaptadas
# =================================================================
@login_required
def crear_directores(request):
    order_by_param = request.GET.get('order_by', '')
    
    if request.method == 'POST':
        formulario = DirectoresForm(request.POST)
        formset = DependenciasFormSet(request.POST)
        
        if formulario.is_valid() and formset.is_valid():
            try:
                # 1. Guardamos el director
                director = formulario.save() 
                
                # 2. Guardamos las dependencias del formset
                instancias = formset.save(commit=False)
                for instancia in instancias:
                    instancia.director = director
                    instancia.save()
                formset.save_m2m()

                # Registro de Log
                LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(director).pk,
                    object_id=director.pk,
                    object_repr=str(director),
                    action_flag=ADDITION,
                    change_message=f'Director {director.cedula} creado con sus dependencias.'
                )
        
                messages.success(request, f'Director {director.cedula} creado exitosamente', extra_tags='procesado ✅')
                return redirect(reverse('listado_directores') + (f'?order_by={order_by_param}' if order_by_param else ''))
            
            except Exception as e:
                logger.error(f"Error al crear Director: {e}")
                messages.error(request, f'Error al guardar: {str(e)}', extra_tags='error ❌')
    else:
        formulario = DirectoresForm()
        formset = DependenciasFormSet()

    return render(request, 'directores/crear.html', {'formulario': formulario, 'formset': formset, 'titulo': 'Nuevo Director'})


#------------------------------------------
@login_required
def editar_directores(request, id):
    director = get_object_or_404(DirectoresBD, id=id) 
    order_by_param = request.GET.get('order_by', '')
    
    if request.method == 'POST':
        formulario = DirectoresForm(request.POST, instance=director)
        formset = DependenciasFormSet(request.POST, instance=director)
        
        if formulario.is_valid() and formset.is_valid():
            try:
                formulario.save()
                formset.save() 

                LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(director).pk,
                    object_id=director.pk,
                    object_repr=str(director),
                    action_flag=CHANGE,
                    change_message=f'Director {director.cedula} actualizado.'
                )

                messages.success(request, f'Director {director.cedula} editado exitosamente', extra_tags='procesado ✅')
                return redirect(reverse('listado_directores') + (f'?order_by={order_by_param}' if order_by_param else ''))

            except Exception as e:
                logger.error(f"Error al editar: {e}")
                messages.error(request, 'Error al guardar los cambios.', extra_tags='error ❌')
    else:
        formulario = DirectoresForm(instance=director)
        formset = DependenciasFormSet(instance=director)

    return render(request, 'directores/editar.html', {'formulario': formulario, 'formset': formset, 'titulo': 'Editar Director'})



#------------------------------------------
@login_required
def borrar_directores(request, id):
# 1. Obtener la instancia del objeto o devolver 404
    # Asume que tu modelo de Director es DirectoresBD
    directores = get_object_or_404(DirectoresBD, id=id) 
    
    # Asume que el código de la Director es 'cedula'
    id_directores = directores.cedula 
    
    # =========================================================
    # 📌 PASO 1: CAPTURAR TODOS LOS PARÁMETROS GET 
    # Esto captura y mantiene el filtro, orden y página.
    # =========================================================
    query_string = request.GET.urlencode()

    try:
        # 2. Intentar la eliminación (puede causar ProtectedError)
        directores.delete()
        
        # 3. Si es exitoso, registrar la acción y mostrar el mensaje de éxito.
        # 🔴 REGISTRO DE LOG (Eliminación Exitosa) --------------------------
        LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(directores).pk,
                    object_id=directores.pk,
                    object_repr=str(directores),
                    action_flag=DELETION, # Bandera para Eliminación (3)
                    change_message=f'Director {id_directores} eliminado mediante vista personalizada.'
                )
        
        messages.success(request, f'Director {id_directores} fue borrado exitosamente', extra_tags='procesado ✅')
        
    except ProtectedError as e:
        # 4. CAPTURAR el ProtectedError (el registro está enlazado)
        logger.error(f"Intento fallido de borrar Director {id_directores}")       
        
        # Generar mensaje de error amigable al usuario
        if e.args and isinstance(e.args[1], set) and e.args[1]:
            # Extraemos el nombre del modelo relacionado que impide la eliminación
            referencing_model = next(iter(e.args[1])).__class__.__name__
            msg = (
                f"ERROR: No se puede eliminar el Director {id_directores}. "
                f"Está enlazada a registros en el modelo **'{referencing_model}'**  "
                f"Por favor, elimine o reasigne los registros enlazados primero."
            )
        else:
            msg = (
                f"ERROR: No se puede eliminar el Director {id_directores} porque "
                f"está siendo referenciada por otros registros. "
                f"Debe eliminarlos o reasignarlos primero."
            )

        messages.error(request, msg, extra_tags='error ❌')
        
    # 5. Redirigir a la lista de directores, ADJUNTANDO todos los parámetros GET.
    # Esto asegura que mantenga el filtro, orden y página.
    base_url = reverse('listado_directores') # Asume que tu URL de listado se llama 'listado_directores'
    
    if query_string:
        # Si hay parámetros (filtros, etc.), los adjuntamos con '?'
        return redirect(f'{base_url}?{query_string}')
    else:
        # Si no había parámetros, redirigimos a la base.
        return redirect('listado_directores')

# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE EXCEL -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_directores_excel(request):
    # --- CAMBIO PARA FILTRO Y ORDEN ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'cedula')

    # 🎯 CAMBIO 1: Se agrega prefetch_related para traer las dependencias y optimizar la consulta
    directores = DirectoresBD.objects.all().prefetch_related('oficinas__dependencia')
    
    if search_query:
        directores = directores.filter(
            Q(cedula__icontains=search_query) | 
            Q(observaciones__icontains=search_query) | 
            Q(usuario__icontains=search_query) | 
            Q(nombres_apellidos__icontains=search_query)
        )
    directores = directores.order_by(order_by)
    # ----------------------------------

    # 1. Configuración Inicial y Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_directores.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte directores"
    
    # --- 2. INSERCIÓN DEL LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        ws.row_dimensions[1].height = 50
        ws.column_dimensions['A'].width = 15 
        img.anchor = 'A1' 
        ws.add_image(img)
    except Exception:
        pass # Silenciar si no encuentra el logo
    
    # --- 3. TÍTULO ---
    ws.merge_cells('B2:E2')
    title_font = Font(name='Arial', size=14, bold=True, color="1F497D") 
    ws['B2'] = "REPORTE DE DIRECTORES Y RESPONSABLES"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # --- 4. ENCABEZADOS ---
    start_row = 4
    # 🎯 CAMBIO 2: Cambiamos el nombre de la columna a 'DEPENDENCIAS ASIGNADAS'
    headers = ['Nro.', 'Código', 'Nombres y Apellidos', 'Dependencias Asignadas', 'Observaciones']
    
    header_font = Font(bold=True, color="FFFFFF") 
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Insertar encabezados
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # --- 5. DATOS DEL MODELO ---
    data_font = Font(size=10, name='Arial')
    center_alignment = Alignment(horizontal='center', vertical='center') 
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    fill_even_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    fill_odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, director in enumerate(directores, 1):
        # 🎯 CAMBIO 3: Lógica para manejar el "Muchos a Muchos"
        # Creamos una lista con los nombres de las dependencias y su cargo
        lista_oficinas = []
        for asignacion in director.oficinas.all():
            texto = f"• {asignacion.dependencia.descripcion} ({asignacion.get_cargo_display()})"
            lista_oficinas.append(texto)
        
        # Unimos todas las dependencias con saltos de línea para que se vean ordenadas en la celda
        dependencias_str = "\n".join(lista_oficinas) if lista_oficinas else "Sin dependencias"

        row_data = [
            i, 
            director.cedula, 
            director.nombres_apellidos.upper(), 
            dependencias_str, 
            director.observaciones
        ]
        
        ws.append(row_data)

        # Aplicar estilos a la fila recién agregada
        current_row_index = start_row + i
        row_fill = fill_even_row if i % 2 == 0 else fill_odd_row
        
        # Ajustar altura de fila si hay múltiples dependencias
        line_count = max(1, len(lista_oficinas))
        ws.row_dimensions[current_row_index].height = 15 * line_count + 10

        for col_index, cell in enumerate(ws[current_row_index]):
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill 
            # Cédula y Nro centrados, el resto a la izquierda
            if col_index in [0, 1]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # --- 6. AUTOAJUSTE DE COLUMNAS ---
    # Definimos anchos fijos razonables para columnas de texto largo
    ws.column_dimensions['A'].width = 6   # Nro
    ws.column_dimensions['B'].width = 15  # Cedula
    ws.column_dimensions['C'].width = 35  # Nombres
    ws.column_dimensions['D'].width = 50  # Dependencias (Más ancho por ser lista)
    ws.column_dimensions['E'].width = 30  # Observaciones
        
    wb.save(response)
    return response


# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE PDF -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_directores_pdf(request):
    
    # --- 1. CAPTURA DE FILTRO Y ORDEN (Mantiene la persistencia) ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'nombres_apellidos')

    try:
        # Iniciamos el queryset
        directores = DirectoresBD.objects.all()
        
        # Aplicamos filtro si existe búsqueda
        if search_query:
            directores = directores.filter(
                Q(cedula__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(usuario__icontains=search_query) | 
                Q(nombres_apellidos__icontains=search_query)
            )
    
        # Aplicamos el orden seleccionado
        directores = directores.order_by(order_by)

    except Exception as e:
        # En caso de error de base de datos o modelo no importado
        print(f"Error en reporte: {e}")
        directores = []

    # --- 2. Preparación del Contexto para la Plantilla PDF ---
    
    # IMPORTANTE: Generar la URL absoluta para el logo
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except AttributeError:
        logo_url = "" 

    # Obtener la fecha y hora actual
    fecha_actual = datetime.now() 
    
    context = {
        'directores': directores, # 👈 Esta variable ya viene filtrada y ordenada
        'titulo_reporte': 'Listado de directores',
        'logo_path': logo_url,
        'fecha_emision': fecha_actual,
    }

    # --- 3. Renderizado y Conversión a PDF ---
    template = get_template('reportes/reporte_directores_pdf.html')
    html = template.render(context)

    result = BytesIO()
    
    pisa_status = pisa.CreatePDF(
       html,              
       dest=result,       
       encoding='utf-8',  
       link_callback=lambda uri, rel: uri
    )

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: %s' % pisa_status.err, status=500)
    
    # --- 4. Devolver la Respuesta con Apertura en Nueva Pestaña ---
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_directores.pdf"'
    
    return response