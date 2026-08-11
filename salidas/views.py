
import os
import csv
import logging
from io import BytesIO
from datetime import datetime
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from xhtml2pdf import pisa

from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.conf import settings
from django.http import HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage
from django.db import IntegrityError, transaction
from django.contrib.contenttypes.models import ContentType
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.db.models import Q, F, Sum, Prefetch  # <--- Agrega Prefetch aquí
from django.contrib.contenttypes.models import ContentType

# Importaciones de modelos y formularios de Salidas
from .models import SalidaEncabezado, SalidaDetalle
from dependencias.models import DependenciasBD
from directores.models import DirectoresBD, DependeciasDirectorBD
from .forms import SalidaEncabezadoForm, SalidaDetalleFormSet, SalidaDetalleFormSet2, SalidaFilterForm # <--- Agrega esto

from django.shortcuts import get_object_or_404
from . import views

logger = logging.getLogger(__name__)

from django.http import JsonResponse
from .models import Producto
from django.views.decorators.http import require_POST

# ----------------------------------------------
@login_required
@require_POST
def revertir_poraprobar_ajax(request):
    # 1. Obtenemos el ID que viene desde el JavaScript
    salida_id = request.POST.get('id')
    
    try:
        # 2. Buscamos La Requisicion usando el nombre real de tu modelo: SalidaEncabezado
        salida = SalidaEncabezado.objects.get(pk=salida_id)
        
        # 3. Cambiamos el estado
        # Según tu models.py, los estados son 'EN_PROCESO', 'POR_APROBAR', 'ANULADA'
        salida.estado = 'EN_PROCESO'
        salida.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'La Requisicion N° {salida.num_requi} ha sido reactivada (EN_PROCESO).'
        })

    except SalidaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró La Requisicion.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ----------------------------------------------
@login_required
@require_POST
def revertir_anulacion_ajax(request):
    # 1. Obtenemos el ID que viene desde el JavaScript
    salida_id = request.POST.get('id')
    
    try:
        # 2. Buscamos La Requisicion usando el nombre real de tu modelo: SalidaEncabezado
        salida = SalidaEncabezado.objects.get(pk=salida_id)
        
        # 3. Cambiamos el estado
        # Según tu models.py, los estados son 'EN_PROCESO', 'POR_APROBAR', 'ANULADA'
        salida.estado = 'EN_PROCESO' 
        salida.save()
        
        return JsonResponse({
            'status': 'success', 
            'message': f'La Requisicion N° {salida.num_requi} ha sido reactivada (EN_PROCESO).'
        })

    except SalidaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró La Requisicion.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ----------------------------------------------
@require_POST
def ejecutar_anulacion_ajax(request):
    # 1. Obtenemos el ID que viene desde el JavaScript
    salida_id = request.POST.get('id')
    
    try:
        # 2. Buscamos La Requisicion usando el nombre real de tu modelo: SalidaEncabezado
        salida = SalidaEncabezado.objects.get(pk=salida_id)
        
        # 3. Cambiamos el estado
        # Según tu models.py, los estados son 'EN_PROCESO', 'POR_APROBAR', 'ANULADA'
        salida.estado = 'ANULADA' 
        salida.save()
        
        return JsonResponse({
            'status': 'success', 
            'message': f'La Requisicion N° {salida.num_requi} ha sido anulada (ANULADA).'
        })
        
    except SalidaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró La Requisicion.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


from django.contrib.auth.decorators import login_required, user_passes_test
@login_required
@user_passes_test(lambda u: u.is_superuser) 
@require_POST
def desbloquear_ajax(request):
    salida_id = request.POST.get('id')
    
    if not salida_id:
        return JsonResponse({'status': 'error', 'message': 'ID no proporcionado'}, status=400)

    try:
        with transaction.atomic():
            # 1. Obtener la Requisición y bloquear la fila para evitar colisiones
            salida = SalidaEncabezado.objects.select_for_update().get(pk=salida_id)
            
            # 2. Validar que el estado sea PROCESADA antes de proceder
            if salida.estado != 'PROCESADA':
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Solo se pueden desbloquear requisiciones PROCESADAS. Estado actual: {salida.estado}'
                })

            # 3. Procesar los detalles para devolver stock
            items = salida.detalles.all() 
            
            if not items.exists():
                 return JsonResponse({'status': 'error', 'message': 'La Requisición no tiene productos registrados.'})

            for item in items:
                prod = item.cod_producto 
                
                # Usamos el campo correcto: cant_entregada
                cantidad_a_devolver = item.cant_entregada if item.cant_entregada else 0
                
                if cantidad_a_devolver > 0:
                    # Sumar de vuelta al maestro de inventario (producto.cantidad)
                    actual_stock = prod.cantidad if prod.cantidad else 0
                    prod.cantidad = actual_stock + cantidad_a_devolver
                    prod.save()
                    
                    # Opcional: Resetear la entrega en el detalle para permitir nueva edición
                    item.cant_entregada = 0
                    item.save()

            # 4. Actualizar el estado del encabezado
            salida.estado = 'POR_APROBAR'
            salida.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': f'Requisición N° {salida.num_requi} desbloqueada. Se han devuelto las cantidades al inventario.'
            })

    except SalidaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Requisición no encontrada.'}, status=404)
    except Exception as e:
        # Si algo falla, el 'transaction.atomic' hace rollback de todo automáticamente
        print(f"ERROR EN DESBLOQUEO: {str(e)}") 
        return JsonResponse({'status': 'error', 'message': f'Error interno: {str(e)}'}, status=500)


@login_required
def listado_salidas(request):
    # 1. Identificar al Director/Usuario y sus dependencias asignadas
    director = DirectoresBD.objects.filter(usuario=request.user.username).first()
    
    # 2. Definir el QuerySet base según el rol
    if request.user.is_superuser or request.user.is_staff:
        queryset_base = SalidaEncabezado.objects.all()
        mis_dependencias = DependenciasBD.objects.all()
    else:
        if director:
            # 🔥 CORRECCIÓN: Filtrar solo relaciones activas
            mis_deps_ids = DependeciasDirectorBD.objects.filter(
                director=director,
                activo=True  # <--- Agregamos este filtro
            ).values_list('dependencia_id', flat=True)
            
            mis_dependencias = DependenciasBD.objects.filter(id__in=mis_deps_ids)
            queryset_base = SalidaEncabezado.objects.filter(cod_dependencia_soli__in=mis_dependencias)
        else:
            queryset_base = SalidaEncabezado.objects.none()
            mis_dependencias = DependenciasBD.objects.none()
            
    # =========================================================
    # 📌 LÓGICA DE FILTRO PERSISTENTE
    # =========================================================
    filtros_datos = request.GET.copy()
    
    # 3. Inicializar el QuerySet con optimización
    queryset = queryset_base.select_related('cod_dependencia_soli').prefetch_related(
        'detalles', 
        'detalles__cod_producto'
    )

    # --- CAMBIO CRÍTICO AQUÍ ---
    # Si es staff y NO hay una búsqueda/filtro activo de estado, restringimos el queryset de entrada
    if (request.user.is_superuser or request.user.is_staff) and 'estado' not in request.GET:
        queryset = queryset.filter(estado__in=['POR_APROBAR', 'EN_PROCESO'])
    # ---------------------------

    # 4. Inicializar el formulario
    filter_form = SalidaFilterForm(filtros_datos or None)
    filter_form.fields['cod_dependencia_soli'].queryset = mis_dependencias

    # 5. Aplicación de Filtros del Formulario
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # Solo aplicamos el filtro de estado si el usuario seleccionó uno explícitamente
        # Esto evita que el "None" del formulario limpie el filtro de arriba
        if data.get('estado'):
            queryset = queryset.filter(estado=data.get('estado'))
        
        # Búsqueda general
        search_query = data.get('search_query')
        if search_query:
            queryset = queryset.filter(
                Q(num_requi__icontains=search_query) | 
                Q(cod_dependencia_soli__descripcion__icontains=search_query)
            ).distinct()

        # Filtros por campos específicos
        if data.get('num_requi'):
            queryset = queryset.filter(num_requi__icontains=data.get('num_requi'))
        
        if data.get('fecha_requi'):
            queryset = queryset.filter(fecha_requi=data.get('fecha_requi'))

        if data.get('fecha_aprobacion'):
            queryset = queryset.filter(fecha_aprobacion=data.get('fecha_aprobacion'))

        if data.get('cod_dependencia_soli'):
            queryset = queryset.filter(cod_dependencia_soli=data.get('cod_dependencia_soli'))

    # 6. Persistencia de filtros
    query_string_with_page = filtros_datos.urlencode()
    
    paginator_params = filtros_datos.copy()
    if 'page' in paginator_params:
        del paginator_params['page']
    query_string_for_paginator = paginator_params.urlencode()

    clean_params = paginator_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    clean_query_string = clean_params.urlencode()

    # 7. Ordenación
    order_by_principal = request.GET.get('order_by', 'num_requi') 
    queryset = queryset.order_by(order_by_principal) 
    
    # 8. Paginación
    page_number = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, Exception):
        page_obj = paginator.page(1)

    # 9. Contexto
    context = {
        'titulo': 'Listado de Salidas / Compras',
        'subpagina': 'subpage',
        'entity': page_obj, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        'query_string': query_string_for_paginator,
        'full_query_string': query_string_with_page,
        'clean_query_string': clean_query_string,
    }
    return render(request, 'salidas/listado.html', context)


# =================================================================
# CREAR SALIDA (ENCABEZADO + FORMSET DE DETALLES)
# =================================================================
@login_required
def crear_salidas(request):
    director = DirectoresBD.objects.filter(usuario=request.user.username).first()
    
    if request.user.is_superuser:
        mis_dependencias = DependenciasBD.objects.all()
    else:
        if director:
            # Filtrar solo relaciones activas
            mis_deps_ids = DependeciasDirectorBD.objects.filter(
                director=director,
                activo=True  # <--- Agregar este filtro
            ).values_list('dependencia_id', flat=True)
            
            mis_dependencias = DependenciasBD.objects.filter(id__in=mis_deps_ids)
        else:
            mis_dependencias = DependenciasBD.objects.none()
    
    # 3. Procesamiento del Formulario
    if request.method == 'POST':
        form = SalidaEncabezadoForm(request.POST)
        
        # IMPORTANTE: Aplicar el filtro de seguridad al queryset del campo antes de validar
        form.fields['cod_dependencia_soli'].queryset = mis_dependencias
        
        if form.is_valid():
            salida = form.save(commit=False)
            salida.estado = 'EN_PROCESO'
            salida.save()
            
            messages.success(
                request, 
                f"✅ ¡Éxito! Se ha grabado la Requisición N° {salida.num_requi}. "
                f"Ahora puedes proceder a cargar los productos."
            )
            return redirect('editar_salidas', id=salida.pk)
    else:
        # Preparación de valores iniciales para el formulario
        hoy = datetime.today().strftime('%Y-%m-%d')
        form = SalidaEncabezadoForm(initial={
            'estado': 'EN_PROCESO',
            'fecha_requi': hoy,
            'fecha_aprobacion': None
        })
        
        # IMPORTANTE: Filtrar las opciones del campo en la carga inicial (GET)
        form.fields['cod_dependencia_soli'].queryset = mis_dependencias
    
    return render(request, 'salidas/crear.html', {
        'formulario': form,
        'titulo': 'Nueva Requisición'
    })


@login_required
@transaction.atomic
def editar_salidas(request, id):
    # 1. OPTIMIZACIÓN: Carga inicial
    encabezado = get_object_or_404(
        SalidaEncabezado.objects.select_related('cod_dependencia_soli').prefetch_related('detalles__cod_producto'), 
        pk=id
    )
  
    # --- LÓGICA DE PERMISOS PARA STAFF/SUPERUSER ---
    es_admin_o_staff = request.user.is_superuser or request.user.is_staff
    # Solo es "lectura" si está por_aprobar Y el usuario NO tiene permisos de staff
    es_solo_lectura = (encabezado.estado != 'EN_PROCESO') and not es_admin_o_staff
    # -----------------------------------------------

    # 2. SEGURIDAD: Identificar al Director y sus dependencias asignadas (SOLO RELACIONES ACTIVAS)
    director = DirectoresBD.objects.filter(usuario=request.user.username).first()
    
    if request.user.is_superuser or request.user.is_staff:
        mis_dependencias = DependenciasBD.objects.all()
    else:
        if director:
            # 🔥 Filtrar solo relaciones activas (mismo filtro que en crear_salidas)
            mis_deps_ids = DependeciasDirectorBD.objects.filter(
                director=director,
                activo=True  # Solo dependencias activas
            ).values_list('dependencia_id', flat=True)
            
            mis_dependencias = DependenciasBD.objects.filter(id__in=mis_deps_ids)
        else:
            mis_dependencias = DependenciasBD.objects.none()
    
        # Verificación de propiedad (Solo si no es superusuario)
        if not request.user.is_superuser and encabezado.cod_dependencia_soli not in mis_dependencias:
            messages.error(request, "No tienes permiso para editar esta Requisición.", extra_tags='error ❌')
            return redirect('listado_salidas')

    query_params = request.GET.copy()
    full_query_string = query_params.urlencode()

    # --- LÓGICA DE FECHA DE HOY (Sincronizada con tu solicitud) ---
    initial_data = {}
    if encabezado.estado == 'POR_APROBAR' and not encabezado.fecha_aprobacion:
        initial_data['fecha_aprobacion'] = datetime.today()

    if request.method == 'POST':
        finalizar_carga = 'btn_finalizar' in request.POST
    
        # Validación de seguridad en el envío del formulario
        if es_solo_lectura:
            messages.error(request, "ERROR: Esta Requisición está cerrada y no tienes permisos para modificarla.", extra_tags='error ❌')
            return redirect('listado_salidas')

        form = SalidaEncabezadoForm(request.POST, instance=encabezado, initial=initial_data)
        if es_admin_o_staff:
            formset = SalidaDetalleFormSet2(request.POST, instance=encabezado)
        else:
            formset = SalidaDetalleFormSet(request.POST, instance=encabezado)

        # Filtramos las dependencias del combo
        form.fields['cod_dependencia_soli'].queryset = mis_dependencias
        finalizar_poraprobar = 'btn_poraprobar' in request.POST
        finalizar_carga = 'btn_finalizar' in request.POST

        if form.is_valid() and formset.is_valid():
            # Validación de duplicados
            productos_vistos = set()
            hay_duplicados = False
            for f in formset.cleaned_data:
                if f and not f.get('DELETE'):
                    prod = f.get('cod_producto')
                    if prod and prod.id in productos_vistos:
                        hay_duplicados = True
                        break
                    if prod: productos_vistos.add(prod.id)

            if hay_duplicados:
                messages.error(request, "No se puede grabar: Hay productos repetidos.", extra_tags='error ❌')
            else:
                try:
                    with transaction.atomic():
                        obj = form.save(commit=False)
                        cambio_a_procesada = False # Variable para disparar el inventario

                        # LÓGICA DE BOTONES:
                        # 1. Si presionan 'btn_poraprobar' (Guardar y solicitar aprobación)
                        if 'btn_poraprobar' in request.POST and obj.estado == 'EN_PROCESO':
                            obj.estado = 'POR_APROBAR'
                        
                        # 2. Si el Administrador presiona 'btn_finalizar' (Aprobar definitivamente)
                        elif 'btn_finalizar' in request.POST and obj.estado == 'POR_APROBAR':
                            obj.estado = 'PROCESADA'
                            cambio_a_procesada = True

                        obj.save()
                        formset.save()

                        # ACTUALIZACIÓN DE INVENTARIO
                        if cambio_a_procesada:
                            # Importante: usar formset.cleaned_data para obtener los objetos finales
                            for f in formset.cleaned_data:
                                if f and not f.get('DELETE'):
                                    producto = f.get('cod_producto')
                                    # Asegúrate que el campo en el Formset sea 'cant_solicitada' o 'cant_entregada'
                                    # según tu modelo. Aquí asumo 'cant_entregada'
                                    cantidad_salida = f.get('cant_entregada') or 0
                                    if cantidad_salida > 0:
                                        Producto.objects.filter(pk=producto.pk).update(
                                            cantidad=F('cantidad') - cantidad_salida
                                        )

                        # Definición del mensaje de auditoría (wilmer)
                        if es_admin_o_staff and encabezado.estado == 'POR_APROBAR' and not cambio_a_procesada:
                            accion = "MODIFICACIÓN ADMINISTRATIVA (CERRADA)"
                        else:
                            accion = "CERRADA" if cambio_a_procesada else "BORRADOR GUARDADO"

                        LogEntry.objects.create(
                            user_id=request.user.pk,
                            content_type_id=ContentType.objects.get_for_model(obj).pk,
                            object_id=obj.pk,
                            object_repr=str(obj),
                            action_flag=CHANGE,
                            change_message=f"Requisicion {obj.num_requi}: {accion}"
                        )

                        messages.success(request, f"Éxito: {accion}", extra_tags='procesado ✅')
                        base_url = reverse('listado_salidas')
                        return redirect(f'{base_url}?{full_query_string}') if full_query_string else redirect(base_url)

                except Exception as e:
                    messages.error(request, f"Error interno: {str(e)}", extra_tags='error ❌')
        else:
            messages.error(request, "Error de validación en los datos.", extra_tags='error ❌')

    else:
        # Método GET: Inicialización normal
        form = SalidaEncabezadoForm(instance=encabezado, initial=initial_data)
        if es_admin_o_staff:
            formset = SalidaDetalleFormSet2(instance=encabezado)
        else:
            formset = SalidaDetalleFormSet(instance=encabezado)
        form.fields['cod_dependencia_soli'].queryset = mis_dependencias

    # Bloqueo visual de campos en el template
    if es_solo_lectura:
        for f in form.fields.values(): f.widget.attrs['disabled'] = 'disabled'
        for sf in formset:
            for f in sf.fields.values(): f.widget.attrs['disabled'] = 'disabled'

    return render(request, 'Salidas/editar.html', {
        'titulo': 'Editar Salida',
        'formulario': form, 
        'formset': formset, 
        'instancia': encabezado,
        'es_solo_lectura': es_solo_lectura,
        'query_string': full_query_string,
    })


#---------------------------------------------------
@login_required
def borrar_salidas(request, id):
    salida = get_object_or_404(SalidaEncabezado, pk=id)
    num_requi = salida.num_requi
    estado_requi = salida.estado
    
    # 📌 CAPTURA DE PARÁMETROS para persistir filtros
    query_params = request.GET.copy()
    full_query_string = query_params.urlencode()

    try:
        with transaction.atomic():

            # Mensaje informativo para la auditoría
            mensaje_auditoria = f'Eliminada Requisicion {estado_requi} N° {num_requi}. No se realizaron ajustes de inventario.'

            # 1. Registrar en el Log de Auditoría antes de borrar
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(salida).pk,
                object_id=salida.pk,
                object_repr=f"Requisicion N° {num_requi}",
                action_flag=DELETION, 
                change_message=mensaje_auditoria
            )
            
            # 2. Eliminar La Requisicion (sus detalles se borrarán por CASCADE si está configurado)
            salida.delete()

        messages.success(request, f'La Requisicion N° {num_requi} ha sido eliminada correctamente.', extra_tags='procesado ✅')
        
    except Exception as e:
        messages.error(request, f'Error al intentar eliminar La Requisicion: {str(e)}', extra_tags='error ❌')

    # 3. Redirección con persistencia de filtros
    base_url = reverse('listado_salidas')
    if full_query_string:
        return redirect(f'{base_url}?{full_query_string}')
    return redirect(base_url)

# =================================================================
# REPORTE PDF ADAPTADO
# =================================================================
@login_required
def reporte_salidas_pdf(request):
    # --- 1. LÓGICA DE SEGURIDAD (FILTRO POR USUARIO/DIRECTOR) ---
    director = DirectoresBD.objects.filter(usuario=request.user.username).first()
    
    if request.user.is_superuser or request.user.is_staff:
        # Administradores ven todas las dependencias
        queryset_base = SalidaEncabezado.objects.all()
    else:
        if director:
            # Usuarios regulares solo ven sus dependencias ACTIVAS
            mis_deps_ids = DependeciasDirectorBD.objects.filter(
                director=director,
                activo=True
            ).values_list('dependencia_id', flat=True)
            queryset_base = SalidaEncabezado.objects.filter(cod_dependencia_soli__in=mis_deps_ids)
        else:
            # Si no hay director asociado, no devuelve nada
            queryset_base = SalidaEncabezado.objects.none()

    # --- 2. OPTIMIZACIÓN Y CAPTURA DE PARÁMETROS ---
    # Importante: Aplicamos optimizaciones sobre el queryset_base ya filtrado
    queryset = queryset_base.select_related('cod_dependencia_soli').prefetch_related(
        'detalles', 
        'detalles__cod_producto'
    )

    # Capturar parámetros del GET
    search_query = request.GET.get('search_query')
    num_requi = request.GET.get('num_requi')
    estado = request.GET.get('estado')
    fecha_requi = request.GET.get('fecha_requi')
    fecha_aprobacion = request.GET.get('fecha_aprobacion')
    cod_dependencia_soli = request.GET.get('cod_dependencia_soli')
    order_by = request.GET.get('order_by', '-fecha_aprobacion')

    # --- 3. APLICACIÓN DE FILTROS ADICIONALES ---
    # Ya no usamos .all() aquí para no romper la seguridad
    if search_query:
        queryset = queryset.filter(
            Q(num_requi__icontains=search_query) | 
            Q(cod_dependencia_soli__descripcion__icontains=search_query)
        ).distinct()

    if num_requi:
        queryset = queryset.filter(num_requi__icontains=num_requi)

    if estado and estado != '':
        queryset = queryset.filter(estado=estado)
    
    if fecha_requi:
        queryset = queryset.filter(fecha_requi=fecha_requi)

    if fecha_aprobacion:
        queryset = queryset.filter(fecha_aprobacion=fecha_aprobacion)

    if cod_dependencia_soli:
        queryset = queryset.filter(cod_dependencia_soli_id=cod_dependencia_soli)

    # --- 4. ORDENACIÓN Y EJECUCIÓN ---
    desempate = 'num_requi' if order_by.lstrip('-') != 'num_requi' else 'fecha_aprobacion'
    salidas = queryset.order_by(order_by, desempate).distinct()

    # --- 5. LÓGICA DE GENERACIÓN DEL PDF ---
    try:
        logo_url = request.build_absolute_uri('/static/img/logo2.png') 
    except AttributeError:
        logo_url = "" 

    context = {
        'entity': salidas,
        'titulo_reporte': 'Reporte de Requisiciones',
        'logo_path': logo_url,
        'fecha_emision': datetime.now(),
    }
    
    template = get_template('reportes/reporte_salidas_pdf.html')
    html = template.render(context)
    result = BytesIO()
    
    # link_callback para que xhtml2pdf gestione las rutas de imágenes
    pisa_status = pisa.CreatePDF(html, dest=result, encoding='utf-8', link_callback=lambda uri, rel: uri)
    
    if pisa_status.err: 
        return HttpResponse('Error al generar PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Reporte_salidas_pdf.pdf"'
    return response


# =================================================================
# REPORTE EXCEL (Exportación Detallada)
# =================================================================
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
@login_required
def reporte_salidas_excel(request):
    # --- 1. LÓGICA DE SEGURIDAD (FILTRO POR USUARIO/DIRECTOR) ---
    director = DirectoresBD.objects.filter(usuario=request.user.username).first()
    
    # Determinamos el universo de datos permitidos para este usuario
    if request.user.is_superuser or request.user.is_staff:
        # Staff y Admin ven todo
        queryset_base = SalidaEncabezado.objects.all()
    else:
        if director:
            # Solo dependencias atribuidas y que estén MARCADAS COMO ACTIVAS
            mis_deps_ids = DependeciasDirectorBD.objects.filter(
                director=director,
                activo=True
            ).values_list('dependencia_id', flat=True)
            
            queryset_base = SalidaEncabezado.objects.filter(cod_dependencia_soli__in=mis_deps_ids)
        else:
            # Si no es staff ni tiene dependencias, no ve nada
            queryset_base = SalidaEncabezado.objects.none()

    # --- 2. OPTIMIZACIÓN Y CAPTURA DE PARÁMETROS ---
    # Aplicamos select_related y prefetch_related sobre el queryset_base filtrado
    queryset = queryset_base.select_related('cod_dependencia_soli').prefetch_related(
        'detalles__cod_producto'
    )

    # Captura de parámetros del GET
    search_query = request.GET.get('search_query')
    num_requi = request.GET.get('num_requi')
    estado = request.GET.get('estado')
    fecha_requi = request.GET.get('fecha_requi')
    fecha_aprobacion = request.GET.get('fecha_aprobacion')
    dependencia_id = request.GET.get('cod_dependencia_soli')
    order_by = request.GET.get('order_by', 'num_requi')

    # --- 3. APLICACIÓN DE FILTROS ADICIONALES ---
    # Importante: No uses .all() aquí para no romper la cadena de filtrado previa
    if search_query:
        queryset = queryset.filter(
            Q(num_requi__icontains=search_query) | 
            Q(cod_dependencia_soli__descripcion__icontains=search_query)
        ).distinct()

    if num_requi:
        queryset = queryset.filter(num_requi__icontains=num_requi)

    if estado and estado != '':
        queryset = queryset.filter(estado=estado)

    if fecha_requi:
        queryset = queryset.filter(fecha_requi=fecha_requi)

    if fecha_aprobacion:
        queryset = queryset.filter(fecha_aprobacion=fecha_aprobacion)

    if dependencia_id:
        # Filtro extra por si el usuario elige una dependencia específica en el buscador
        queryset = queryset.filter(cod_dependencia_soli_id=dependencia_id)

    # Orden final
    salidas = queryset.order_by(order_by)
    
    # --- Generación de Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Requisiciones"

    # --- 2. INSERCIÓN DEL LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo2.png')
    try:
        img = OpenpyxlImage(logo_path)
        ws.row_dimensions[1].height = 50 
        ws.column_dimensions['A'].width = 15 
        img.anchor = 'A1' 
        ws.add_image(img)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el logo en la ruta: {logo_path}")


    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A4D94", end_color="1A4D94", fill_type="solid") # Azul del PDF
    center_aligned = Alignment(horizontal='center', vertical='center')
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Título ---
    ws.merge_cells('A1:J1')
    ws['A1'] = "REPORTE DE REQUISICIONES"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_aligned

    # --- Encabezados de la Tabla ---
    headers = [
        'Requisición', 'Fecha Requi.', 'F. Aprobación', 'Dependencia', 
        'Estado', 'Insumo', 'Descripción', 'U. Medida', 'Cant. Solicitada', 'Cant. Entregada'
    ]
    
    start_row = 3
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = border_style

    # --- Cuerpo de los Datos ---
    current_row = start_row + 1
    for salida in salidas:
        detalles = salida.detalles.all()
        
        if not detalles.exists():
            # Si no hay productos, mostrar fila vacía con info del encabezado
            ws.cell(row=current_row, column=1).value = salida.num_requi
            ws.cell(row=current_row, column=2).value = salida.fecha_requi.strftime('%d/%m/%Y') if salida.fecha_requi else ""
            ws.cell(row=current_row, column=3).value = salida.fecha_aprobacion.strftime('%d/%m/%Y') if salida.fecha_aprobacion else "Pendiente"
            ws.cell(row=current_row, column=4).value = str(salida.cod_dependencia_soli.descripcion).upper() if salida.cod_dependencia_soli else "S/D"
            ws.cell(row=current_row, column=5).value = salida.estado
            ws.cell(row=current_row, column=6).value = "---"
            ws.cell(row=current_row, column=7).value = "REQUISICIÓN SIN PRODUCTOS"
            current_row += 1
        else:
            for detalle in detalles:
                ws.cell(row=current_row, column=1).value = salida.num_requi
                ws.cell(row=current_row, column=2).value = salida.fecha_requi.strftime('%d/%m/%Y') if salida.fecha_requi else ""
                ws.cell(row=current_row, column=3).value = salida.fecha_aprobacion.strftime('%d/%m/%Y') if salida.fecha_aprobacion else "Pendiente"
                ws.cell(row=current_row, column=4).value = str(salida.cod_dependencia_soli.descripcion).upper() if salida.cod_dependencia_soli else "S/D"
                ws.cell(row=current_row, column=5).value = salida.estado
                
                # Datos del producto
                ws.cell(row=current_row, column=6).value = detalle.cod_producto.cod_producto if detalle.cod_producto else ""
                ws.cell(row=current_row, column=7).value = detalle.cod_producto.descripcion if detalle.cod_producto else ""
                ws.cell(row=current_row, column=8).value = str(detalle.cod_producto.cod_unidad) if detalle.cod_producto else ""
                ws.cell(row=current_row, column=9).value = float(detalle.cant_solicitada or 0)
                ws.cell(row=current_row, column=10).value = float(detalle.cant_entregada or 0)
                
                # Formato numérico para cantidad
                ws.cell(row=current_row, column=9).number_format = '#,##0.00'
                ws.cell(row=current_row, column=10).number_format = '#,##0.00'
                current_row += 1

    # Ajuste automático de columnas
    column_widths = [12, 12, 12, 35, 15, 12, 45, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Retorno del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Requisiciones_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# --------------------------------------
@login_required
def imprimir_requisicion(request, id):
    # 1. Obtener la salida
    salida = get_object_or_404(
        SalidaEncabezado.objects.prefetch_related('detalles__cod_producto'), 
        num_requi=id
    )
    
    # 2. ENLACE CRÍTICO: Buscar el director usando el campo cod_dependencia_soli
    # Buscamos en la tabla intermedia el director que pertenece a esa dependencia
    relacion = DependeciasDirectorBD.objects.filter(
        dependencia=salida.cod_dependencia_soli
    ).select_related('director').first()

    # Si existe la relación, extraemos el nombre del director
    nombre_solicitante = "No asignado"
    if relacion and relacion.director:
        nombre_solicitante = relacion.director.nombres_apellidos # O el campo que guarde el nombre en DirectoresBD

    # 3. NUEVO: Buscar a la persona encargada de Operatividad
    # Primero buscamos la dependencia
    dep_operatividad = DependenciasBD.objects.filter(
        descripcion__icontains="OPERATIVIDAD Y FUNCIONAMIENTO"
    ).first()

    nombre_autorizador = "No asignado"
    if dep_operatividad:
        # Buscamos la relación activa en la tabla intermedia para esa dependencia específica
        relacion_autorizador = DependeciasDirectorBD.objects.filter(
            dependencia=dep_operatividad,
            activo=True
        ).select_related('director').first()
        
        if relacion_autorizador:
            nombre_autorizador = relacion_autorizador.director.nombres_apellidos

    # 4. NUEVO: Buscar a la persona encargada de del almacen
    # Primero buscamos la dependencia
    dep_almacen = DependenciasBD.objects.filter(
        descripcion__icontains="ALMACEN OPERATIVIDAD"
    ).first()

    nombre_entregador = "No asignado"
    if dep_almacen:
        # Buscamos la relación activa en la tabla intermedia para esa dependencia específica
        relacion_entregador = DependeciasDirectorBD.objects.filter(
            dependencia=dep_almacen,
            activo=True
        ).select_related('director').first()
        
        if relacion_entregador:
            nombre_entregador = relacion_entregador.director.nombres_apellidos

    context = {
        'salida': salida,
        'nombre_solicitante': nombre_solicitante, # Pasamos el nombre al PDF
        'autorizado_por': nombre_autorizador,  # <--- Pasamos el objeto completo
        'entregado_por': nombre_entregador,  # <--- Pasamos el objeto completo
        'logo_path': request.build_absolute_uri(settings.STATIC_URL + 'img/logo_inven.png') if settings.STATIC_URL else "",
        'fecha_emision': datetime.now(),
        'titulo_reporte': 'Requisiciones',
    }

    # 4. Cargar el template y renderizarlo con el contexto
    # Asegúrate de que la ruta al archivo .html sea la correcta en tu carpeta templates
    template = get_template('reportes/reporte_requisicion_pdf.html')
    html = template.render(context)
    
    # 5. Crear el buffer para el PDF
    result = BytesIO()
    
    # Generar el PDF
    pisa_status = pisa.CreatePDF(
        html, 
        dest=result, 
        encoding='utf-8'
    )

    # 6. Verificar errores
    if pisa_status.err:
        return HttpResponse('Error técnico al generar el reporte PDF', status=500)
        
    # 7. Construir la respuesta del navegador
    filename = f"Requisicion_{salida.num_requi}.pdf"
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    
    # 'inline' hace que se abra en el navegador, 'attachment' obligaría a descargar
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    return response