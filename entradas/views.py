
import os
import csv
import logging
from io import BytesIO
from datetime import datetime
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from xhtml2pdf import pisa

from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.conf import settings
from django.http import HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage
from django.db import IntegrityError, transaction
from django.contrib.contenttypes.models import ContentType
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.db.models import Q, F, Sum, Prefetch  # <--- Agrega Prefetch aquí
from django.contrib.contenttypes.models import ContentType

# Importaciones de modelos y formularios de Entradas
from .models import EntradaEncabezado, EntradaDetalle
from .forms import EntradaEncabezadoForm, EntradaDetalleFormSet, EntradaFilterForm # <--- Agrega esto

from django.shortcuts import get_object_or_404
from . import views

logger = logging.getLogger(__name__)

from django.http import JsonResponse
from .models import Producto
from django.views.decorators.http import require_POST

# ----------------------------------------------
@login_required
@require_POST
def revertir_anulacion_entrada_ajax(request):
    # 1. Obtenemos el ID que viene desde el JavaScript
    entrada_id = request.POST.get('id')
    
    try:
        # 2. Buscamos la orden usando el nombre real de tu modelo: EntradaEncabezado
        entrada = EntradaEncabezado.objects.get(pk=entrada_id)
        
        # 3. Cambiamos el estado
        # Según tu models.py, los estados son 'EN_PROCESO', 'PROCESADA', 'ANULADA'
        entrada.estado = 'EN_PROCESO' 
        entrada.save()
        
        return JsonResponse({
            'status': 'success', 
            'message': f'La orden N° {entrada.num_orden} ha sido reactivada (EN_PROCESO).'
        })
        
    except EntradaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró la orden.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ----------------------------------------------
@login_required
@require_POST
def ejecutar_anulacion_entrada_ajax(request):
    entrada_id = request.POST.get('id')
    try:
        entrada = EntradaEncabezado.objects.get(pk=entrada_id)
        
        # Validación de seguridad: Solo permitir anular si está EN_PROCESO
        # Si está PROCESADA, debería usarse la función de "desbloquear" primero.
        if entrada.estado != 'EN_PROCESO':
            return JsonResponse({
                'status': 'error', 
                'message': f'No se puede anular una orden con estado {entrada.estado}.'
            }, status=400)
        entrada.estado = 'ANULADA'
        entrada.save()
        
        return JsonResponse({
            'status': 'success', 
            'message': f'La orden N° {entrada.num_orden} ha sido anulada.'
        })
        
    except EntradaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró la orden.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

from django.contrib.auth.decorators import login_required, user_passes_test
@login_required
@user_passes_test(lambda u: u.is_superuser) 
@require_POST
def desbloquear_entrada_ajax(request):
    entrada_id = request.POST.get('id')
    
    if not entrada_id:
        return JsonResponse({'status': 'error', 'message': 'ID no proporcionado'}, status=400)

    try:
        with transaction.atomic():
            # 1. Obtener la orden usando tu modelo real
            entrada = EntradaEncabezado.objects.get(pk=entrada_id)
            
            # 2. Verificar estado para evitar errores
            if entrada.estado == 'EN_PROCESO':
                return JsonResponse({'status': 'error', 'message': 'La orden ya está en proceso.'})

            # 3. Procesar detalles usando el related_name='detalles'
            items = entrada.detalles.all() 
            
            if not items.exists():
                 return JsonResponse({'status': 'error', 'message': 'La orden no tiene productos registrados.'})

            for item in items:
                # Usamos 'cod_producto' que es el nombre en tu modelo EntradaDetalle
                prod = item.cod_producto 
                
                # Validamos valores nulos para evitar errores matemáticos
                actual_stock = prod.cantidad if prod.cantidad else 0
                cant_recibida = item.cant_recibida if item.cant_recibida else 0
                
                # Restamos la cantidad del maestro de inventario
                prod.cantidad = max(0, actual_stock - cant_recibida)
                prod.save()

            # 4. Cambiamos el estado a EN_PROCESO
            entrada.estado = 'EN_PROCESO'
            entrada.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': f'Orden N° {entrada.num_orden} desbloqueada. Se descontó la cantidad del inventario.'
            })
            
    except EntradaEncabezado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Orden no encontrada.'}, status=404)
    except Exception as e:
        # Imprime el error real en tu terminal para debug
        print(f"ERROR EN DESBLOQUEO: {str(e)}") 
        return JsonResponse({'status': 'error', 'message': f'Error interno: {str(e)}'}, status=500)

# =================================================================
# VISTA PRINCIPAL: LISTADO DE ENTRADAS (COMPRAS)
# =================================================================
@login_required
def listado_entradas(request):
    # 📌 1. Definir el estado inicial si no hay parámetros en la URL
    if not request.GET:
        filter_form = EntradaFilterForm(request.GET)
        queryset_base = EntradaEncabezado.objects.all()
    else:
        filter_form = EntradaFilterForm(request.GET)
        queryset_base = EntradaEncabezado.objects.all()

    # QuerySet con optimización
    queryset = queryset_base.select_related('ced_proveedor').prefetch_related(
        'detalles', 
        'detalles__cod_producto'
    )

    # =========================================================
    # 📌 PERSISTENCIA (Sin cambios, se mantiene igual)
    # =========================================================
    full_query_params = request.GET.copy()
    query_string_with_page = full_query_params.urlencode()
    
    paginator_params = full_query_params.copy()
    if 'page' in paginator_params:
        del paginator_params['page']
    query_string_for_paginator = paginator_params.urlencode()

    clean_params = paginator_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    clean_query_string = clean_params.urlencode()

    # =========================================================

    # 3. Procesamiento de filtros
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        
        # Filtro General
        search_query = data.get('search_query')
        if search_query:
            queryset = queryset.filter(
                Q(num_orden__icontains=search_query) | 
                Q(num_factura__icontains=search_query) |
                Q(ced_proveedor__nombres_apellidos__icontains=search_query) |
                Q(ced_proveedor__ced_proveedor__icontains=search_query)
            ).distinct()

        # Filtros Específicos
        if data.get('num_orden'):
            queryset = queryset.filter(num_orden__icontains=data.get('num_orden'))
        
        # 📌 Si el usuario seleccionó un estado manualmente, se aplica aquí
        if data.get('estado'):
            queryset = queryset.filter(estado=data.get('estado'))
            
        if data.get('fec_orden'):
            queryset = queryset.filter(fecha_orden=data.get('fec_orden'))

        if data.get('num_factura'):
            queryset = queryset.filter(num_factura__icontains=data.get('num_factura'))

        if data.get('fec_factura'):
            queryset = queryset.filter(fecha_factura=data.get('fec_factura'))

        if data.get('ced_proveedor'):
            queryset = queryset.filter(ced_proveedor=data.get('ced_proveedor'))

    # 4. Ordenación
    order_by_principal = request.GET.get('order_by', 'num_orden') 
    campos_ordenar = [order_by_principal]
    desempate = 'num_orden' if order_by_principal.lstrip('-') != 'num_orden' else 'fecha_orden'
    campos_ordenar.append(desempate)
    queryset = queryset.order_by(*campos_ordenar) 
    
    # 5. Paginación (Se mantiene igual)
    page_number = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)
    try:
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except:
        raise Http404('*** Página no encontrada ***')

    # 6. Contexto
    context = {
        'titulo': 'Listado de Entradas / Compras',
        'subpagina': 'subpage',
        'entity': page_obj, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        'query_string': query_string_for_paginator,
        'full_query_string': query_string_with_page,
        'clean_query_string': clean_query_string,
    }
    return render(request, 'entradas/listado.html', context)

# =================================================================
# CREAR ENTRADA (ENCABEZADO + FORMSET DE DETALLES)
# =================================================================
from datetime import date # <--- Asegúrate de tener esta importación
@login_required
def crear_entradas(request):
    if request.method == 'POST':
        form = EntradaEncabezadoForm(request.POST)
        if form.is_valid():
            entrada = form.save(commit=False)
            entrada.estado = 'EN_PROCESO' 
            entrada.save()
            # MODIFICACIÓN: Mensaje con el número de orden grabado
            messages.success(
                request, 
                f"✅ ¡Éxito! Se ha grabado la Orden N° {entrada.num_orden}. "
                f"Ahora puedes proceder a cargar los productos."
            )
            return redirect('editar_entradas', id=entrada.pk)
    else:
        # Obtenemos la fecha de hoy en formato ISO (YYYY-MM-DD)
        hoy = date.today().strftime('%Y-%m-%d')
        
        # Añadimos las fechas al diccionario initial
        form = EntradaEncabezadoForm(initial={
            'estado': 'EN_PROCESO',
            'fecha_orden': hoy,    # <--- Fecha del día para la Orden
            'fecha_factura': hoy    # <--- Fecha del día para la Factura
        })
    
    return render(request, 'entradas/crear.html', {
        'formulario': form,
        'titulo': 'Nueva Orden de Compra'
    })

@login_required
@transaction.atomic
def editar_entradas(request, id):
    encabezado = get_object_or_404(EntradaEncabezado, pk=id)
    query_params = request.GET.copy()
    full_query_string = query_params.urlencode()
    es_solo_lectura = encabezado.estado == 'PROCESADA'

    if request.method == 'POST':

        if es_solo_lectura:
            messages.error(request, "ERROR: Esta orden ya está cerrada.", extra_tags='error ❌')
            return redirect('listado_entradas')

        form = EntradaEncabezadoForm(request.POST, instance=encabezado)
        formset = EntradaDetalleFormSet(request.POST, instance=encabezado)
        
        # Identificamos si se presionó el botón de finalizar
        finalizar_carga = 'btn_finalizar' in request.POST

        if form.is_valid() and formset.is_valid():
            # --- VALIDACIÓN DE DUPLICADOS ---
            productos_vistos = set()
            hay_duplicados = False
            for f in formset.cleaned_data:
                if f and not f.get('DELETE'):
                    prod = f.get('cod_producto')
                    if prod and prod.id in productos_vistos:
                        hay_duplicados = True
                        break
                    if prod: productos_vistos.add(prod.id)

            if hay_duplicados:
                messages.error(request, "No se puede grabar: Hay productos repetidos.", extra_tags='error ❌')
            else:
                try:
                    with transaction.atomic():
                        # A. GUARDAR CABECERA
                        obj = form.save(commit=False)

                        cambio_a_procesada = False
                        if finalizar_carga and obj.estado == 'EN_PROCESO':
                            obj.estado = 'PROCESADA'
                            cambio_a_procesada = True
                        obj.save()

                        # B. GUARDAR DETALLES (Esto unifica el borrador y la carga final)
                        formset.save()

                        # C. AUMENTO DE INVENTARIO
                        if cambio_a_procesada:
                            # Usamos los datos ya validados del formset para actualizar el stock
                            for f in formset.cleaned_data:
                                if f and not f.get('DELETE'):
                                    producto = f.get('cod_producto')
                                    cantidad_nueva = f.get('cant_recibida') or 0
                                    
                                    if cantidad_nueva > 0:
                                        # Actualización directa para evitar condiciones de carrera
                                        Producto.objects.filter(pk=producto.pk).update(
                                            cantidad=F('cantidad') + cantidad_nueva
                                        )

                        # D. AUDITORÍA Y MENSAJE
                        accion = "CERRADA" if cambio_a_procesada else "BORRADOR GUARDADO"
                        LogEntry.objects.create(
                            user_id=request.user.pk,
                            content_type_id=ContentType.objects.get_for_model(obj).pk,
                            object_id=obj.pk,
                            object_repr=str(obj),
                            action_flag=CHANGE,
                            change_message=f"Orden {obj.num_orden}: {accion}"
                        )

                        messages.success(request, f"Éxito: {accion}", extra_tags='procesado ✅')
                        base_url = reverse('listado_entradas')
                        return redirect(f'{base_url}?{full_query_string}') if full_query_string else redirect(base_url)

                except Exception as e:
                    messages.error(request, f"Error: {str(e)}", extra_tags='error ❌')
        else:
            # Mostrar errores detallados en consola para debug
            print("Errores Formset:", formset.errors)
            messages.error(request, "Error de validación en los datos enviados.", extra_tags='error ❌')

    else:
        form = EntradaEncabezadoForm(instance=encabezado)
        formset = EntradaDetalleFormSet(instance=encabezado)

    # Bloqueo de campos si ya está procesada (Vista)
    if es_solo_lectura:
        for f in form.fields.values(): f.widget.attrs['disabled'] = 'disabled'
        for sf in formset:
            for f in sf.fields.values(): f.widget.attrs['disabled'] = 'disabled'

    return render(request, 'entradas/editar.html', {
        'titulo': 'Editar Entrada',
        'formulario': form, 
        'formset': formset, 
        'instancia': encabezado,
        'es_solo_lectura': es_solo_lectura,
        'query_string': full_query_string,
    })


#---------------------------------------------------
@login_required
def borrar_entradas(request, id):
    entrada = get_object_or_404(EntradaEncabezado, pk=id)
    num_orden = entrada.num_orden
    estado_orden = entrada.estado
    
    # 📌 CAPTURA DE PARÁMETROS para persistir filtros
    query_params = request.GET.copy()
    full_query_string = query_params.urlencode()

    try:
        with transaction.atomic():

            # Mensaje informativo para la auditoría
            mensaje_auditoria = f'Eliminada Orden {estado_orden} N° {num_orden}. No se realizaron ajustes de inventario.'

            # 1. Registrar en el Log de Auditoría antes de borrar
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(entrada).pk,
                object_id=entrada.pk,
                object_repr=f"Orden de Entrada N° {num_orden}",
                action_flag=DELETION, 
                change_message=mensaje_auditoria
            )
            
            # 2. Eliminar la orden (sus detalles se borrarán por CASCADE si está configurado)
            entrada.delete()

        messages.success(request, f'La Orden N° {num_orden} ha sido eliminada correctamente.', extra_tags='procesado ✅')
        
    except Exception as e:
        messages.error(request, f'Error al intentar eliminar la orden: {str(e)}', extra_tags='error ❌')

    # 3. Redirección con persistencia de filtros
    base_url = reverse('listado_entradas')
    if full_query_string:
        return redirect(f'{base_url}?{full_query_string}')
    return redirect(base_url)

# =================================================================
# REPORTE PDF ADAPTADO
# =================================================================
@login_required
def reporte_entradas_pdf(request):
    # 1. Capturar parámetros
    search_query = request.GET.get('search_query')
    num_orden = request.GET.get('num_orden')
    estado = request.GET.get('estado')
    fec_orden = request.GET.get('fec_orden')
    num_factura = request.GET.get('num_factura')
    fec_factura = request.GET.get('fec_factura')
    ced_proveedor = request.GET.get('ced_proveedor')
    order_by = request.GET.get('order_by', '-fecha_factura')

    # 2. QuerySet Base con optimización
    queryset = EntradaEncabezado.objects.select_related('ced_proveedor').prefetch_related(
        'detalles', 
        'detalles__cod_producto'
    ).all()

    # 3. Aplicar Filtros
    if search_query:
        queryset = queryset.filter(
            Q(num_orden__icontains=search_query) | 
            Q(num_factura__icontains=search_query) |
            Q(ced_proveedor__nombres_apellidos__icontains=search_query)
        ).distinct()

    if num_orden:
        queryset = queryset.filter(num_orden__icontains=num_orden)

    if estado and estado != '':
        queryset = queryset.filter(estado=estado)
    
    if fec_orden:
        queryset = queryset.filter(fecha_orden=fec_orden)

    if num_factura:
        queryset = queryset.filter(num_factura__icontains=num_factura)

    if fec_factura:
        queryset = queryset.filter(fecha_factura=fec_factura)

    if ced_proveedor:
        queryset = queryset.filter(ced_proveedor_id=ced_proveedor)

    # 4. Ordenación y ejecución
    desempate = 'num_orden' if order_by.lstrip('-') != 'num_orden' else 'fecha_factura'
    entradas = queryset.order_by(order_by, desempate).distinct()

    # --- NUEVO: Lógica para el Logo (Copiado de Productos) ---
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except AttributeError:
        logo_url = "" 

    # 5. Generación del PDF
    context = {
        'entity': entradas,
        'titulo_reporte': 'Reporte de Reposiciones (Ordenes Compras)',
        'logo_path': logo_url, # <--- Se agrega al contexto
        'fecha_emision': datetime.now(),
    }
    
    template = get_template('reportes/reporte_entradas_pdf.html')
    html = template.render(context)
    result = BytesIO()
    
    # Se agrega el link_callback para que xhtml2pdf encuentre la imagen
    pisa_status = pisa.CreatePDF(html, dest=result, encoding='utf-8', link_callback=lambda uri, rel: uri)
    
    if pisa_status.err: return HttpResponse('Error al generar PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Reporte_Entradas_pdf.pdf"'
    return response


# --------------------------------------
@login_required
def imprimir_entradas(request, id):
    # 1. Obtener la instancia usando num_orden (que viene como 'id' desde la URL)
    # Usamos prefetch_related para cargar los detalles y productos en una sola consulta
    entrada = get_object_or_404(
        EntradaEncabezado.objects.prefetch_related('detalles__cod_producto'), 
        num_orden=id
    )
    
    # 2. Configurar la ruta del logo para el PDF
    # Intentamos construir una ruta absoluta para que pisa pueda encontrar la imagen
    try:
        logo_url = request.build_absolute_uri(settings.STATIC_URL + 'img/logo_inven.png')
    except Exception:
        logo_url = ""

    # 3. Preparar el contexto que usará el template HTML
    context = {
        'entrada': entrada,             # Objeto principal
        'logo_path': logo_url,          # Ruta del logo
        'fecha_emision': datetime.now(),
        'titulo_reporte': 'Orden de Entrada de Inventario',
    }

    # 4. Cargar el template y renderizarlo con el contexto
    # Asegúrate de que la ruta al archivo .html sea la correcta en tu carpeta templates
    template = get_template('reportes/reporte_orden_pdf.html')
    html = template.render(context)
    
    # 5. Crear el buffer para el PDF
    result = BytesIO()
    
    # Generar el PDF
    pisa_status = pisa.CreatePDF(
        html, 
        dest=result, 
        encoding='utf-8'
    )

    # 6. Verificar errores
    if pisa_status.err:
        return HttpResponse('Error técnico al generar el reporte PDF', status=500)
        
    # 7. Construir la respuesta del navegador
    filename = f"Orden_Entrada_{entrada.num_orden}.pdf"
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    
    # 'inline' hace que se abra en el navegador, 'attachment' obligaría a descargar
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    return response


# =================================================================
# REPORTE EXCEL (Sincronizado con Campos Específicos) - SUBSANADO
# =================================================================
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@login_required
def reporte_entradas_excel(request):
    # 1. Captura de parámetros
    search_query = request.GET.get('search_query')
    num_orden = request.GET.get('num_orden')
    estado = request.GET.get('estado')
    fec_orden = request.GET.get('fec_orden')
    num_factura = request.GET.get('num_factura')
    fec_factura = request.GET.get('fec_factura')
    proveedor_id = request.GET.get('proveedor') or request.GET.get('ced_proveedor')
    order_by = request.GET.get('order_by', '-fecha_factura')

    # 2. QuerySet Base
    queryset = EntradaEncabezado.objects.select_related('ced_proveedor').all()

    # 3. Aplicación de Filtros
    if search_query:
        queryset = queryset.filter(
            Q(num_orden__icontains=search_query) | 
            Q(num_factura__icontains=search_query) |
            Q(ced_proveedor__nombres_apellidos__icontains=search_query)
        )
    if num_orden: queryset = queryset.filter(num_orden__icontains=num_orden)
    if estado: queryset = queryset.filter(estado=estado)
    if fec_orden: queryset = queryset.filter(fecha_orden=fec_orden)
    if num_factura: queryset = queryset.filter(num_factura__icontains=num_factura)
    if fec_factura: queryset = queryset.filter(fecha_factura=fec_factura)
    if proveedor_id: queryset = queryset.filter(ced_proveedor_id=proveedor_id)

    # 4. Prefetch de detalles optimizado para traer producto y su unidad
    detalles_prefetch = Prefetch(
        'detalles',
        # Ajusta 'cod_unidad' si el campo en el modelo Producto tiene otro nombre
        queryset=EntradaDetalle.objects.select_related('cod_producto', 'cod_producto__cod_unidad'),
        to_attr='detalles_filtrados'
    )
    entradas = queryset.prefetch_related(detalles_prefetch).order_by(order_by).distinct()

    # --- Generación de Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Entradas"

    # --- Estilos Globales ---
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Inserción del Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        img.width, img.height = 100, 50
        ws.add_image(img, 'A1')
    except: pass

    # --- Título ---
    ws.merge_cells('B3:I3') # Expandido a la columna I
    ws['B3'] = "REPORTE DE ENTRADAS DE INVENTARIO"
    ws['B3'].font = Font(name='Arial', size=14, bold=True, color="1F497D")
    ws['B3'].alignment = Alignment(horizontal='center')

    # --- Encabezados (9 columnas en total) ---
    headers = [
        'ORDEN', 'ESTADO', 'FACTURA', 'PROVEEDOR', 'MONTO FACT.', 
        'CODIGO', 'PRODUCTO', 'UNIDAD', 'CANT. REC.'
    ]
    
    start_row = 5
    header_fill = PatternFill(start_color="3B18D6", end_color="3B18D6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # --- Cuerpo del Reporte ---
    current_row = start_row + 1
    
    for entrada in entradas:
        detalles = getattr(entrada, 'detalles_filtrados', [])
        num_detalles = len(detalles)
        row_inicio = current_row
        
        status = entrada.estado
        fecha_o = entrada.fecha_orden.strftime('%d/%m/%Y') if entrada.fecha_orden else ""
        proveedor = f"{entrada.ced_proveedor.nombres_apellidos}\n{entrada.ced_proveedor.ced_proveedor}" if entrada.ced_proveedor else "No Registrado"
        monto_f = float(entrada.monto_factura or 0)

        if num_detalles == 0:
            # Caso sin productos: se llenan las primeras 5 columnas y se une el resto
            ws.cell(row=current_row, column=1).value = f"{entrada.num_orden}\n{fecha_o}"
            ws.cell(row=current_row, column=2).value = status
            ws.cell(row=current_row, column=3).value = f"{entrada.num_factura or '---'}\n{entrada.fecha_factura or ''}"
            ws.cell(row=current_row, column=4).value = proveedor
            ws.cell(row=current_row, column=5).value = monto_f
            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
            
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=9)
            cell_msg = ws.cell(row=current_row, column=6)
            cell_msg.value = "Orden sin productos registrados"
            cell_msg.font = Font(italic=True, color="808080")
            cell_msg.alignment = Alignment(horizontal='center')
            
            for c in range(1, 10): ws.cell(row=current_row, column=c).border = border_thin
            current_row += 1
        else:
            # Caso con productos
            for det in detalles:
                # 6: CODIGO
                ws.cell(row=current_row, column=6).value = det.cod_producto.cod_producto.upper() if det.cod_producto else "S/C"
                # 7: PRODUCTO (Descripción)
                ws.cell(row=current_row, column=7).value = det.cod_producto.descripcion.upper() if det.cod_producto else "S/D"
                # 8: UNIDAD (Traída desde la relación del producto)
                try:
                    # Se asume que cod_unidad es una relación y queremos su descripción/nombre
                    ws.cell(row=current_row, column=8).value = det.cod_producto.cod_unidad.descripcion.upper()
                except:
                    # Si cod_unidad es un CharField o falla la relación
                    ws.cell(row=current_row, column=8).value = str(getattr(det.cod_producto, 'cod_unidad', 'N/A')).upper()
                
                # 9: CANT. REC.
                ws.cell(row=current_row, column=9).value = float(det.cant_recibida or 0)
                ws.cell(row=current_row, column=9).number_format = '#,##0.00'
                ws.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')
                
                for c in range(1, 10): ws.cell(row=current_row, column=c).border = border_thin
                current_row += 1

            # Agrupación de celdas de encabezado (Rowspan) de la columna 1 a la 5
            for col in range(1, 6):
                ws.merge_cells(start_row=row_inicio, start_column=col, end_row=current_row-1, end_column=col)
                cell = ws.cell(row=row_inicio, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if col == 1: cell.value = f"{entrada.num_orden}\n{fecha_o}"
                elif col == 2: cell.value = status
                elif col == 3: cell.value = f"{entrada.num_factura or '---'}\n{entrada.fecha_factura or ''}"
                elif col == 4: cell.value = proveedor
                elif col == 5: 
                    cell.value = monto_f
                    cell.number_format = '#,##0.00'

    # Ajuste de anchos de columna para 9 columnas
    column_widths = [12, 15, 15, 35, 15, 15, 40, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Entradas.xlsx"'
    wb.save(response)
    return response
