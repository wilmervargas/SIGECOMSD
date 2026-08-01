import os
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from django.conf import settings
from django.views.decorators.http import require_POST # Opcional, pero bueno para APIs
from django.views.decorators.csrf import csrf_exempt

# SEGURIDAD HISTORICO DE LOS REGISTROS
from django.contrib.contenttypes.models import ContentType # Para identificar el modelo
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION      # LogEntry y Flags de acción
# 💡 NUEVA IMPORTACIÓN PARA MANEJO DE ERRORES DE BASE DE DATOS
from django.db import IntegrityError 
# 💡 NUEVA IMPORTACIÓN PARA EL REGISTRO DE EVENTOS (LOGGING)
import logging
from django.db.models import ProtectedError # 👈 Importar el error específico
# -------------------------------------------------------------

# Importa el formulario de unidad
from unidades.forms import UnidadFilterForm, UnidadForm
from unidades.models import UnidadBD

from django.db.models import Q, F, Sum # Importamos Sum
from django.db.models import Prefetch

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage
from django.http import Http404, HttpResponse
from django.urls import reverse
from datetime import datetime

logger = logging.getLogger(__name__)

@login_required
def listado_unidad(request):
    # 1. Inicialización
    filter_form = UnidadFilterForm(request.GET)
    queryset_unidad = UnidadBD.objects.all() 

    # =========================================================
    # 📌 PERSISTENCIA TOTAL: Captura filtros, orden Y página
    # =========================================================
    
    # 1. Obtenemos todos los parámetros actuales
    full_query_params = request.GET.copy()
    
    # 2. query_string_with_page: Para los botones "Editar" y "Borrar"
    query_string_with_page = full_query_params.urlencode()
    
    # 3. query_string (sin página): Para los enlaces del PAGINADOR
    paginator_params = full_query_params.copy()
    if 'page' in paginator_params:
        del paginator_params['page']
    query_string_for_paginator = paginator_params.urlencode()

    # 4. clean_query_string: Solo filtros (sin orden ni página) para headers
    clean_params = paginator_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    clean_query_string = clean_params.urlencode()
    # =========================================================

    # 2. Lógica de Filtrado
    if filter_form.is_valid():
        search_query = filter_form.cleaned_data.get('search_query')
        if search_query:
            queryset_unidad = queryset_unidad.filter(
                Q(cod_unidad__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )

    # 3. Lógica de Ordenación
    order_by_principal = request.GET.get('order_by', 'cod_unidad') 
    if order_by_principal.lstrip('-') == 'descripcion':
        lista_final = queryset_unidad.order_by(order_by_principal, 'cod_unidad')
    else:
        lista_final = queryset_unidad.order_by(order_by_principal)

    # 4. Paginación
    page_number = request.GET.get('page', 1)
    paginator = Paginator(lista_final, 50)
    try:
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except:
        # Captura errores como strings en el parámetro 'page'
        raise Http404('*** Página no encontrada ***')

    # 5. Contexto
    datos = {
        'titulo': 'Unidades',
        'subpagina': 'subpage', 
        'entity': page_obj, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        
        # Variables de persistencia para el HTML:
        'query_string': query_string_for_paginator, 
        'full_query_string': query_string_with_page, 
        'clean_query_string': clean_query_string, 
    }
    
    return render(request, 'unidades/listado.html', datos)

# =================================================================
# VISTAS CRUD (Crear, Editar, Borrar) - Variables Adaptadas
# =================================================================
@login_required
def crear_unidad(request):

    formulario = UnidadForm(request.POST or None) 
    datos = {'titulo': 'Tabla Unidades', 'subpagina': 'subpage', 'formulario': formulario,}
    order_by_param = request.GET.get('order_by', '')
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Contiene el código que debe ser monitoreado
            
            # 1. Guarda y obtén el objeto Producto creado
            unidad = formulario.save() 
            
            # 🟢 REGISTRO DE LOG (Creación Exitosa) --------------------------
            LogEntry.objects.create(
                        user_id=request.user.pk,
                        content_type_id=ContentType.objects.get_for_model(unidad).pk,
                        object_id=unidad.pk,
                        object_repr=str(unidad),
                        action_flag=DELETION, # Bandera para Eliminación (3)
                        change_message=f'Unidad {unidad.cod_unidad} creada mediante vista personalizada.'
                    )
            # --------------------------------------------------------

            messages.success(request, 'Unidad creada exitosamente', 
                             extra_tags='procesado ✅')
            base_url = reverse('listado_unidad')
            return redirect(f'{base_url}?order_by={order_by_param}') if order_by_param else redirect('listado_unidad')
        
        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores (Overflow, DB o cualquier otro)
            
            # 1. Registro detallado en el Log del Sistema (Recomendado)
            # exc_info=True asegura que se grabe el StackTrace (la pila de llamadas)
            logger.error(f"Fallo crítico al crear Unidad para usuario {request.user.pk}: {e}", exc_info=True)
            
            # 2. Registro de Fallo en la tabla LogEntry (Para el historial del Admin)
            # Usamos 4 (o el valor que definas para ERROR) ya que ADDITION, CHANGE y DELETION son 1, 2 y 3.
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(UnidadBD).pk, 
                object_id=None, # No hay ID de objeto porque la creación falló
                object_repr='Fallo en la creación de Unidad', 
                action_flag=4, # Usamos '4' (o un número fuera de ADDITION/CHANGE/DELETION) para Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo crear Unidad. Causa: {type(e).__name__}. Ver log del servidor.'
            )

            # 3. Muestra un mensaje amigable al usuario
            messages.error(request, 'Error grave de ejecución. La Unidad NO fue creada. Por favor, contacte a soporte.', extra_tags='error ❌')

    return render (request, 'unidades/crear.html', datos)


#------------------------------------------
@login_required
@login_required
def editar_unidad(request, id):
    unidad = get_object_or_404(UnidadBD, id=id) 
    
    # 📌 CAPTURAMOS TODOS LOS PARÁMETROS (Filtro, Orden y Página) para persistencia
    query_params = request.GET.copy()
    query_string = query_params.urlencode()

    formulario = UnidadForm(request.POST or None, instance=unidad)
    
    datos = {
        'titulo': 'Tabla Unidades', 
        'subpagina': 'subpage', 
        'formulario': formulario, 
        'query_string': query_string, # 👈 Para que el botón "Cancelar" también mantenga el estado
    }
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Monitoreo de ejecución
            
            # 1. Capturar campos modificados
            campos_modificados = formulario.changed_data 
            
            # 2. Guardar cambios
            unidad = formulario.save() 
            
            # 3. Construir mensaje de cambio para el Log
            if campos_modificados:
                change_message = 'Campos modificados: ' + ', '.join(campos_modificados)
            else:
                change_message = 'La unidad fue guardada, pero no se detectaron cambios.'
                
            # 🟡 REGISTRO DE LOG (Modificación Exitosa) ------------------------
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(unidad).pk,
                object_id=unidad.pk,
                object_repr=str(unidad),
                action_flag=CHANGE, # ✅ Corregido a CHANGE (edición)
                change_message=f'Unidad {unidad.cod_unidad} editada mediante vista personalizada. {change_message}'
            )
            # --------------------------------------------------------

            messages.success(request, f'Unidad {unidad.cod_unidad} editada exitosamente', extra_tags='procesado ✅')
            
            # 📌 REDIRECCIÓN: Mantenemos filtros, orden y página original
            base_url = reverse('listado_unidad')
            if query_string:
                return redirect(f'{base_url}?{query_string}#pagtable')
            return redirect('listado_unidad')

        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Registro de fallos
            logger.error(f"Fallo crítico al editar unidad ID {id} para usuario {request.user.pk}: {e}", exc_info=True)
            
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(UnidadBD).pk, 
                object_id=None,
                object_repr='Fallo en la edición de Unidad',
                action_flag=4, # Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo editar Unidad. Causa: {type(e).__name__}. Ver log del servidor.'
            )

            messages.error(request, 'Error grave de ejecución al editar. Los cambios NO fueron guardados. Contacte a soporte.', extra_tags='error ❌')

    return render(request, 'unidades/editar.html', datos)

#------------------------------------------
@login_required
def borrar_unidad(request, id):
# 1. Obtener la instancia del objeto o devolver 404
    # Asegúrate de importar el modelo UnidadBD al inicio de views.py
    unidad = get_object_or_404(UnidadBD, id=id) 
    
    # Asume que el código de la unidad es 'cod_unidad'
    id_unidad = unidad.cod_unidad 
    
    # =========================================================
    # 📌 PASO 1: CAPTURAR TODOS LOS PARÁMETROS GET 
    # Mantiene filtro, orden y página.
    # =========================================================
    query_string = request.GET.urlencode()

    try:
        # 2. Intentar la eliminación (puede causar ProtectedError)
        unidad.delete()
        
        # 3. Si es exitoso, registrar la acción y mostrar el mensaje de éxito.
        # 🔴 REGISTRO DE LOG (Eliminación Exitosa) --------------------------
        LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(unidad).pk,
                    object_id=unidad.pk,
                    object_repr=str(unidad),
                    action_flag=DELETION, # Bandera para Eliminación (3)
                    change_message=f'Unidad {id_unidad} eliminada mediante vista personalizada.'
                )
        # ----------------------------------------------------------
        
        messages.success(request, f'Unidad {id_unidad} fue borrada exitosamente', extra_tags='procesado ✅')
        
    except ProtectedError as e:
        # 4. CAPTURAR el ProtectedError (el registro está enlazado)
        logger.error(f"Intento fallido de borrar Unidad {id_unidad}")       
        
        # Generar mensaje de error amigable al usuario
        if e.args and isinstance(e.args[1], set) and e.args[1]:
            # Extraemos el nombre del modelo relacionado que impide la eliminación
            referencing_model = next(iter(e.args[1])).__class__.__name__
            msg = (
                f"ERROR: No se puede eliminar la unidad {id_unidad}. "
                f"Está enlazada a registros en el modelo **'{referencing_model}'**  "
                f"Por favor, elimine o reasigne los registros enlazados primero."
            )
        else:
            msg = (
                f"ERROR: No se puede eliminar la unidad {id_unidad} porque "
                f"está siendo referenciada por otros registros. "
                f"Debe eliminarlos o reasignarlos primero."
            )

        messages.error(request, msg, extra_tags='error ❌')
        
    # 5. Redirigir a la lista de unidades, ADJUNTANDO todos los parámetros GET.
    # Esto asegura que mantenga el filtro, orden y página.
    base_url = reverse('listado_unidad') # Asume que tu URL de listado se llama 'listado_unidad'
    
    if query_string:
        # Si hay parámetros (filtros, etc.), los adjuntamos con '?'
        return redirect(f'{base_url}?{query_string}')
    else:
        # Si no había parámetros, redirigimos a la base.
        return redirect('listado_unidad')

# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE EXCEL UNIDADES ------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_unidades_excel(request):

    # --- CAMBIO PARA FILTRO Y ORDEN ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'cod_unidad')

    unidades = UnidadBD.objects.all()
    if search_query:
        unidades = unidades.filter(
            Q(cod_unidad__icontains=search_query) | 
            Q(observaciones__icontains=search_query) |
            Q(descripcion__icontains=search_query)
        )
    unidades = unidades.order_by(order_by)
    # ----------------------------------

    # 1. Configuración Inicial y Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Unidades_Formato.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Unidades"
    
    # --- 2. INSERCIÓN DEL LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        ws.row_dimensions[1].height = 50 
        ws.column_dimensions['A'].width = 15 
        img.anchor = 'A1' 
        ws.add_image(img)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el logo en la ruta: {logo_path}")
    
    # --- 3. CREACIÓN Y ESTILOS DEL ENCABEZADO ---
    ws.merge_cells('B2:D2')
    title_font = Font(name='Arial', size=14, bold=True, color="1F497D") 
    title_alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'] = "REPORTE DE UNIDADES DE MEDIDA"
    ws['B2'].font = title_font
    ws['B2'].alignment = title_alignment
    
    # --- 4. ENCABEZADOS DE DATOS ---
    ws.append(()) 
    header_font = Font(bold=True, color="FFFFFF") 
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    data_font = Font(size=10, name='Arial')
    center_alignment = Alignment(horizontal='center', vertical='center') 
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    
    fill_even_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    fill_odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    start_row = 4
    headers = ['Nro.', 'Código', 'Descripción Unidad', 'Observaciones']
    ws.append(headers) 

    for cell in ws[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
        
    # --- 5. DATOS DEL MODELO ---
    data_start_row = start_row + 1 
    
    for i, uni in enumerate(unidades, 1): 
        row_data = [i, uni.cod_unidad, uni.descripcion, uni.observaciones]
        ws.append(row_data)

        current_row_index = data_start_row + i - 1
        current_row = ws[current_row_index]
        ws.row_dimensions[current_row_index].height = 40
        row_fill = fill_even_row if i % 2 == 0 else fill_odd_row
        
        for col_index, cell in enumerate(current_row):
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill 
            if col_index in [0, 1]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # --- 6. AUTOAJUSTE ---
    column_widths = {} 
    for row in ws.iter_rows(min_row=start_row):
        for i, cell in enumerate(row):
            col_letter = get_column_letter(i + 1)
            try:
                if cell.value is not None:
                    length = max(10, len(str(cell.value))) 
                    current_max = column_widths.get(col_letter, 0)
                    if length > current_max: column_widths[col_letter] = length
            except: pass 

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width + 3 
        
    wb.save(response)
    return response

# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE PDF UNIDADES --------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_unidades_pdf(request):
    # --- 1. CAPTURA DE FILTRO Y ORDEN ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'descripcion')

    try:
        unidades = UnidadBD.objects.all()
        if search_query:
            unidades = unidades.filter(
                Q(cod_unidad__icontains=search_query) | 
                Q(observaciones__icontains=search_query) |
                Q(descripcion__icontains=search_query)
            )
        unidades = unidades.order_by(order_by)
    except Exception as e:
        print(f"Error en reporte unidades: {e}")
        unidades = []

    # --- 2. Preparación del Contexto ---
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except AttributeError:
        logo_url = "" 

    fecha_actual = datetime.now() 
    
    context = {
        'unidades': unidades, 
        'titulo_reporte': 'Listado de Unidades de Medida',
        'logo_path': logo_url,
        'fecha_emision': fecha_actual,
    }

    # --- 3. Renderizado y Conversión ---
    # Nota: Asegúrate de tener creada esta plantilla similar a la de categorías
    template = get_template('reportes/reporte_unidades_pdf.html')
    html = template.render(context)

    result = BytesIO()
    pisa_status = pisa.CreatePDF(
       html,              
       dest=result,       
       encoding='utf-8',  
       link_callback=lambda uri, rel: uri
    )

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: %s' % pisa_status.err, status=500)
    
    # --- 4. Devolver Respuesta ---
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_unidades.pdf"'
    
    return response