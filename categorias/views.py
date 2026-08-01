
import os
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa
from django.conf import settings
from django.views.decorators.http import require_POST # Opcional, pero bueno para APIs
from django.views.decorators.csrf import csrf_exempt

# SEGURIDAD HISTORICO DE LOS REGISTROS
from django.contrib.contenttypes.models import ContentType # Para identificar el modelo
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION      # LogEntry y Flags de acción
# 💡 NUEVA IMPORTACIÓN PARA MANEJO DE ERRORES DE BASE DE DATOS
from django.db import IntegrityError 
# 💡 NUEVA IMPORTACIÓN PARA EL REGISTRO DE EVENTOS (LOGGING)
import logging
from django.db.models import ProtectedError # 👈 Importar el error específico

# -------------------------------------------------------------

# Importa el formulario de Categoría
from categorias.forms import CategoriaFilterForm, CategoriaForm
from categorias.models import CategoriaBD

from django.db.models import Q, F, Sum # Importamos Sum
from django.db.models import Prefetch

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage
from django.http import Http404, HttpResponse
from django.urls import reverse
from datetime import datetime

logger = logging.getLogger(__name__)

@login_required
def listado_categoria(request):
    filter_form = CategoriaFilterForm(request.GET)
    queryset_categoria = CategoriaBD.objects.all() 

    # =========================================================
    # 📌 PERSISTENCIA TOTAL: Captura filtros, orden Y página
    # =========================================================
    
    # 1. Obtenemos todos los parámetros actuales (search_query, order_by, page, etc.)
    full_query_params = request.GET.copy()
    
    # 2. query_string_with_page: Se usará para los botones "Editar" y "Borrar"
    # Esto garantiza que al volver de editar, regreses a la misma página.
    query_string_with_page = full_query_params.urlencode()
    
    # 3. query_string (sin página): Se usa para los enlaces del PAGINADOR
    # (Tu HTML ya añade ?page=..., por lo que aquí debemos quitarlo para no duplicar)
    paginator_params = full_query_params.copy()
    if 'page' in paginator_params:
        del paginator_params['page']
    query_string_for_paginator = paginator_params.urlencode()

    # 4. clean_query_string: Solo filtros (sin orden ni página) para los headers de la tabla
    clean_params = paginator_params.copy()
    if 'order_by' in clean_params:
        del clean_params['order_by']
    clean_query_string = clean_params.urlencode()

    # =========================================================

    # Lógica de Filtrado
    if filter_form.is_valid():
        search_query = filter_form.cleaned_data.get('search_query')
        if search_query:
            queryset_categoria = queryset_categoria.filter(
                Q(cod_categoria__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query))

    # Lógica de Ordenación
    order_by_principal = request.GET.get('order_by', 'cod_categoria') 
    if order_by_principal.lstrip('-') == 'descripcion':
        lista_categoria = queryset_categoria.order_by(order_by_principal, 'cod_categoria')
    else:
        lista_categoria = queryset_categoria.order_by(order_by_principal)

    # --- Paginación ---
    page_number = request.GET.get('page', 1)
    paginator = Paginator(lista_categoria, 50)
    try:
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except:
        raise Http404('*** Página no encontrada ***')

    # --- Contexto ---
    datos = {
        'titulo': 'Categorías',
        'subpagina': 'subpage', 
        'entity': page_obj, 
        'paginator': paginator,
        'order_by': order_by_principal,
        'filter_form': filter_form,
        
        # Estas variables son las que usará tu HTML:
        'query_string': query_string_for_paginator, # Para los números del paginador
        'full_query_string': query_string_with_page, # PARA LOS BOTONES EDITAR/BORRAR
        'clean_query_string': clean_query_string, 
    }
    return render(request, 'categorias/listado.html', datos)

# =================================================================
# VISTAS CRUD (Crear, Editar, Borrar) - Variables Adaptadas
# =================================================================
@login_required
def crear_categoria(request):

    formulario = CategoriaForm(request.POST or None) 
    datos = {'titulo': 'Tabla Categorías', 'subpagina': 'subpage', 'formulario': formulario,}
    order_by_param = request.GET.get('order_by', '')
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Contiene el código que debe ser monitoreado
            
            # 1. Guarda y obtén el objeto Producto creado
            categoria = formulario.save() 
            
            # 🟢 REGISTRO DE LOG (Creación Exitosa) --------------------------
            LogEntry.objects.create(
                        user_id=request.user.pk,
                        content_type_id=ContentType.objects.get_for_model(categoria).pk,
                        object_id=categoria.pk,
                        object_repr=str(categoria),
                        action_flag=DELETION, # Bandera para Eliminación (3)
                        change_message=f'Categoría {categoria.cod_categoria} creada mediante vista personalizada.'
                    )
        
            messages.success(request, f'Categoría {categoria.cod_categoria} fue creada exitosamente', extra_tags='procesado ✅')

            base_url = reverse('listado_categoria')
            return redirect(f'{base_url}?order_by={order_by_param}') if order_by_param else redirect('listado_categoria')
        
        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores (Overflow, DB o cualquier otro)
            
            # 1. Registro detallado en el Log del Sistema (Recomendado)
            # exc_info=True asegura que se grabe el StackTrace (la pila de llamadas)
            logger.error(f"Fallo crítico al crear categoría para usuario {request.user.pk}: {e}", exc_info=True)
            
            # 2. Registro de Fallo en la tabla LogEntry (Para el historial del Admin)
            # Usamos 4 (o el valor que definas para ERROR) ya que ADDITION, CHANGE y DELETION son 1, 2 y 3.
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(CategoriaBD).pk, 
                object_id=None, # No hay ID de objeto porque la creación falló
                object_repr='Fallo en la creación de categoría', 
                action_flag=4, # Usamos '4' (o un número fuera de ADDITION/CHANGE/DELETION) para Fallo/Error
                change_message=f'ERROR CRÍTICO: No se pudo crear la categoría. Causa: {type(e).__name__}. Ver log del servidor.'
            )
            # 3. Muestra un mensaje amigable al usuario
            messages.error(request, 'Error grave de ejecución. La categoría NO fue creada. Por favor, contacte a soporte.', extra_tags='error ❌')

    return render (request, 'categorias/crear.html', datos)


#------------------------------------------
@login_required
def editar_categoria(request, id):
#------------------------------------------
    categoria = get_object_or_404(CategoriaBD, id=id) 
    
    # 📌 CAPTURAMOS TODOS LOS PARÁMETROS (Filtro, Orden y Página)
    query_params = request.GET.copy()
    query_string = query_params.urlencode()

    formulario = CategoriaForm(request.POST or None, instance=categoria)
    
    datos = {
        'titulo': 'Tabla Categorías', 
        'subpagina': 'subpage', 
        'formulario': formulario, 
        'query_string': query_string, # 👈 Pasamos todo el string a la plantilla
    }
    
    if formulario.is_valid():
        try:
            # 🎯 BLOQUE TRY: Contiene el código que debe ser monitoreado
            
            # 1. Capturar la lista de campos modificados antes de guardar
            campos_modificados = formulario.changed_data 
            
            # 2. Guardar los cambios
            categoria = formulario.save() 
            
            # 3. Construir el mensaje de cambio
            if campos_modificados:
                change_message = 'Campos modificados: ' + ', '.join(campos_modificados)
            else:
                change_message = 'categoría fue guardada, pero no se detectaron cambios.'
                
            # 🟡 REGISTRO DE LOG (Modificación Exitosa) ------------------------
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(categoria).pk,
                object_id=categoria.pk,
                object_repr=str(categoria),
                action_flag=CHANGE, # 💡 Nota: CHANGE es la bandera correcta para edición
                change_message=f'Categoría {categoria.cod_categoria} editada mediante vista personalizada. {change_message}'
            )
            # --------------------------------------------------------

            messages.success(request, f'Categoría {categoria.cod_categoria} editada exitosamente', extra_tags='procesado ✅')
            
            # 📌 REDIRECCIÓN: Si hay parámetros (página, filtro, orden), los incluimos todos
            base_url = reverse('listado_categoria')
            if query_string:
                return redirect(f'{base_url}?{query_string}#pagtable')
            return redirect('listado_categoria')

        except (OverflowError, IntegrityError, Exception) as e:
            # 🛑 BLOQUE EXCEPT: Captura errores durante la edición
            logger.error(f"Fallo crítico al editar catgoría ID {id} para usuario {request.user.pk}: {e}", exc_info=True)
            
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(CategoriaBD).pk, 
                object_id=None,
                object_repr='Fallo en la edición de categoría',
                action_flag=4,
                change_message=f'ERROR CRÍTICO: No se pudo editar la categoría. Causa: {type(e).__name__}. Ver log del servidor.'
            )

            messages.error(request, 'Error grave de ejecución al editar. Los cambios NO fueron guardados. Contacte a soporte.', extra_tags='error ❌')

    return render(request, 'categorias/editar.html', datos)

#------------------------------------------
@login_required
def borrar_categoria(request, id):
# 1. Obtener la instancia del objeto o devolver 404
    # Asume que tu modelo de Categoría es CategoriaBD
    categoria = get_object_or_404(CategoriaBD, id=id) 
    
    # Asume que el código de la categoría es 'cod_categoria'
    id_categoria = categoria.cod_categoria 
    
    # =========================================================
    # 📌 PASO 1: CAPTURAR TODOS LOS PARÁMETROS GET 
    # Esto captura y mantiene el filtro, orden y página.
    # =========================================================
    query_string = request.GET.urlencode()

    try:
        # 2. Intentar la eliminación (puede causar ProtectedError)
        categoria.delete()
        
        # 3. Si es exitoso, registrar la acción y mostrar el mensaje de éxito.
        # 🔴 REGISTRO DE LOG (Eliminación Exitosa) --------------------------
        LogEntry.objects.create(
                    user_id=request.user.pk,
                    content_type_id=ContentType.objects.get_for_model(categoria).pk,
                    object_id=categoria.pk,
                    object_repr=str(categoria),
                    action_flag=DELETION, # Bandera para Eliminación (3)
                    change_message=f'Categoría {id_categoria} eliminada mediante vista personalizada.'
                )
        
        messages.success(request, f'Categoría {id_categoria} fue borrada exitosamente', extra_tags='procesado ✅')
        
    except ProtectedError as e:
        # 4. CAPTURAR el ProtectedError (el registro está enlazado)
        logger.error(f"Intento fallido de borrar Categoría {id_categoria}")       
        
        # Generar mensaje de error amigable al usuario
        if e.args and isinstance(e.args[1], set) and e.args[1]:
            # Extraemos el nombre del modelo relacionado que impide la eliminación
            referencing_model = next(iter(e.args[1])).__class__.__name__
            msg = (
                f"ERROR: No se puede eliminar la categoría {id_categoria}. "
                f"Está enlazada a registros en el modelo **'{referencing_model}'**  "
                f"Por favor, elimine o reasigne los registros enlazados primero."
            )
        else:
            msg = (
                f"ERROR: No se puede eliminar la categoría {id_categoria} porque "
                f"está siendo referenciada por otros registros. "
                f"Debe eliminarlos o reasignarlos primero."
            )

        messages.error(request, msg, extra_tags='error ❌')
        
    # 5. Redirigir a la lista de categorías, ADJUNTANDO todos los parámetros GET.
    # Esto asegura que mantenga el filtro, orden y página.
    base_url = reverse('listado_categoria') # Asume que tu URL de listado se llama 'listado_categoria'
    
    if query_string:
        # Si hay parámetros (filtros, etc.), los adjuntamos con '?'
        return redirect(f'{base_url}?{query_string}')
    else:
        # Si no había parámetros, redirigimos a la base.
        return redirect('listado_categoria')

# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE EXCEL -----------------------------------------
# ---------------------------------------------------------------------------------
@login_required
def reporte_categorias_excel(request):

    # --- CAMBIO PARA FILTRO Y ORDEN ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'cod_categoria')

    categorias = CategoriaBD.objects.all()
    if search_query:
        categorias = categorias.filter(
            Q(cod_categoria__icontains=search_query) | 
            Q(observaciones__icontains=search_query) | 
            Q(descripcion__icontains=search_query)
        )
    categorias = categorias.order_by(order_by)
    # ----------------------------------

    # 1. Configuración Inicial y Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Categorias_Formato.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Categorias"
    
    # --- 2. INSERCIÓN DEL LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_inven.png')
    try:
        img = OpenpyxlImage(logo_path)
        ws.row_dimensions[1].height = 50
        ws.column_dimensions['A'].width = 15 
        img.anchor = 'A1' 
        ws.add_image(img)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el logo en la ruta: {logo_path}")
    
    # --- 3. CREACIÓN Y ESTILOS DEL ENCABEZADO (Título) ---
    ws.merge_cells('B2:D2')
    title_font = Font(name='Arial', size=14, bold=True, color="1F497D") 
    title_alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'] = "REPORTE DE CATEGORIAS"
    ws['B2'].font = title_font
    ws['B2'].alignment = title_alignment
    
    # --- 4. ENCABEZADOS DE DATOS Y ESTILOS DE TABLA ---
    ws.append(()) 
    header_font = Font(bold=True, color="FFFFFF") 
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    data_font = Font(size=10, name='Arial')
    center_alignment = Alignment(horizontal='center', vertical='center') 
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    
    fill_even_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    fill_odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    start_row = 4
    headers = ['Nro.', 'Código', 'Descripción Categoría', 'Observaciones']
    ws.append(headers) 

    for cell in ws[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # --- 5. DATOS DEL MODELO ---
    data_start_row = start_row + 1 
    
    # SE ELIMINÓ LA LÍNEA QUE SOBREESCRIBÍA LAS CATEGORÍAS FILTRADAS
    for i, categ in enumerate(categorias, 1): 
        row_data = [i, categ.cod_categoria, categ.descripcion, categ.observaciones]
        ws.append(row_data)

        current_row_index = data_start_row + i - 1
        current_row = ws[current_row_index]
        ws.row_dimensions[current_row_index].height = 40
        row_fill = fill_even_row if i % 2 == 0 else fill_odd_row
        
        for col_index, cell in enumerate(current_row):
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill 
            if col_index in [0, 1]:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # --- 6. AUTOAJUSTE Y LIMITACIÓN DE COLUMNAS ---
    column_widths = {} 
    MAX_WIDTH_C = 40 
    MAX_WIDTH_D = 40 

    for row in ws.iter_rows(min_row=start_row):
        for i, cell in enumerate(row):
            col_letter = get_column_letter(i + 1)
            try:
                if cell.value is not None:
                    length = max(10, len(str(cell.value))) 
                    if i == 2: length = min(length, MAX_WIDTH_C)
                    elif i == 3: length = min(length, MAX_WIDTH_D)
                    current_max = column_widths.get(col_letter, 0)
                    if length > current_max:
                        column_widths[col_letter] = length
            except:
                pass 

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width + 3 
        
    wb.save(response)
    return response


# ---------------------------------------------------------------------------------
# -------------------- ESTE ES REPORTE PDF -----------------------------------------
# ---------------------------------------------------------------------------------

@login_required
def reporte_categorias_pdf(request):
    # --- 1. CAPTURA DE FILTRO Y ORDEN (Mantiene la persistencia) ---
    search_query = request.GET.get('search_query')
    order_by = request.GET.get('order_by', 'descripcion')

    try:
        # Iniciamos el queryset
        categorias = CategoriaBD.objects.all()
        
        # Aplicamos filtro si existe búsqueda
        if search_query:
            categorias = categorias.filter(
                Q(cod_categoria__icontains=search_query) | 
                Q(observaciones__icontains=search_query) | 
                Q(descripcion__icontains=search_query)
            )
        
        # Aplicamos el orden seleccionado
        categorias = categorias.order_by(order_by)
        
    except Exception as e:
        # En caso de error de base de datos o modelo no importado
        print(f"Error en reporte: {e}")
        categorias = []

    # --- 2. Preparación del Contexto para la Plantilla PDF ---
    
    # IMPORTANTE: Generar la URL absoluta para el logo
    try:
        logo_url = request.build_absolute_uri('/static/img/logo_inven.png') 
    except AttributeError:
        logo_url = "" 

    # Obtener la fecha y hora actual
    fecha_actual = datetime.now() 
    
    context = {
        'categorias': categorias, # 👈 Esta variable ya viene filtrada y ordenada
        'titulo_reporte': 'Listado de Categorías',
        'logo_path': logo_url,
        'fecha_emision': fecha_actual,
    }

    # --- 3. Renderizado y Conversión a PDF ---
    template = get_template('reportes/reporte_categorias_pdf.html')
    html = template.render(context)

    result = BytesIO()
    
    pisa_status = pisa.CreatePDF(
       html,              
       dest=result,       
       encoding='utf-8',  
       link_callback=lambda uri, rel: uri
    )

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: %s' % pisa_status.err, status=500)
    
    # --- 4. Devolver la Respuesta con Apertura en Nueva Pestaña ---
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_categorias.pdf"'
    
    return response